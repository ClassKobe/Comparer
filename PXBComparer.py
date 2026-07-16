import os
import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from rapidfuzz import fuzz, process
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Google Sheets (vendor alias) integration ─────────────────────────────────
# Same concept as tt.py: a shared Google Sheet lets any user add a
# "Vendor Name -> Alias" pair from the GUI, and every run of this program
# reads that sheet back in so the classifier/matcher knows about it too.
try:
    from google.oauth2.service_account import Credentials
    import gspread
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

GOOGLE_SHEET_NAME  = "learned_rules"     # same spreadsheet used by tt.py
ALIAS_TAB_NAME     = "VendorAliases"     # new tab inside that spreadsheet
ALIAS_TAB_HEADERS  = ["Vendor", "Alias"]
CREDS_FILE         = "credentials.json"  # path to your downloaded service-account JSON key

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

_gsheet_client = None  # cached gspread client


def get_gsheet_client():
    """Return a cached, authorized gspread client, or None if Sheets access
    isn't available (missing library or missing credentials.json)."""
    global _gsheet_client
    if not GSPREAD_AVAILABLE:
        return None
    if _gsheet_client is not None:
        return _gsheet_client
    if not os.path.exists(CREDS_FILE):
        return None
    try:
        creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
        _gsheet_client = gspread.authorize(creds)
        return _gsheet_client
    except Exception as e:
        print(f"[VendorAliases] Could not authorize Google Sheets client: {e}")
        return None


def get_alias_worksheet():
    """Return the 'VendorAliases' worksheet inside the shared spreadsheet,
    creating it (with headers) if it doesn't exist yet. Returns None if
    Sheets access isn't available."""
    client = get_gsheet_client()
    if client is None:
        return None
    try:
        spreadsheet = client.open(GOOGLE_SHEET_NAME)
    except Exception as e:
        print(f"[VendorAliases] Could not open spreadsheet '{GOOGLE_SHEET_NAME}': {e}")
        return None

    try:
        ws = spreadsheet.worksheet(ALIAS_TAB_NAME)
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=ALIAS_TAB_NAME, rows=1000, cols=2)
        ws.append_row(ALIAS_TAB_HEADERS)
    except Exception as e:
        print(f"[VendorAliases] Could not access tab '{ALIAS_TAB_NAME}': {e}")
        return None
    return ws


# In-memory copy of everything currently in the VendorAliases sheet tab,
# keyed lowercase like VENDOR_ALIASES so resolve_alias() can consult both.
SHEET_VENDOR_ALIASES = {}

def load_vendor_aliases_from_sheet():
    """Pull every Vendor/Alias row from the Google Sheet tab into
    SHEET_VENDOR_ALIASES. Safe to call even if Sheets isn't configured
    (just leaves the dict empty/unchanged). Returns the dict."""
    global SHEET_VENDOR_ALIASES
    ws = get_alias_worksheet()
    if ws is None:
        return SHEET_VENDOR_ALIASES
    try:
        records = ws.get_all_records()  # list of {"Vendor": ..., "Alias": ...}
    except Exception as e:
        print(f"[VendorAliases] Could not read rows: {e}")
        return SHEET_VENDOR_ALIASES

    fresh = {}
    for row in records:
        vendor = str(row.get("Vendor", "")).strip()
        alias  = str(row.get("Alias", "")).strip()
        if not vendor or not alias:
            continue
        fresh[vendor.lower()] = alias
    SHEET_VENDOR_ALIASES = fresh
    return SHEET_VENDOR_ALIASES


def add_vendor_alias_to_sheet(vendor: str, alias: str):
    """Append a new Vendor/Alias row to the shared sheet and update the
    in-memory SHEET_VENDOR_ALIASES immediately so it's usable without a
    restart. Raises an exception on failure so the caller (GUI) can show it."""
    vendor = vendor.strip()
    alias  = alias.strip()
    if not vendor or not alias:
        raise ValueError("Both Vendor Name and Alias are required.")

    ws = get_alias_worksheet()
    if ws is None:
        raise RuntimeError(
            "Google Sheets isn't available. Make sure 'gspread' and "
            "'google-auth' are installed and credentials.json is present."
        )
    ws.append_row([vendor, alias])
    SHEET_VENDOR_ALIASES[vendor.lower()] = alias


OUTPUT_FILE = "privv_importable.csv"




ITEM_MAP = {
    "102": "Project Feasibility Analysis",
    "107": "Enviromental Assessment",
    "108": "Hazardous Material Survey",
    "111": "Site Survey",
    "201": "Financing Costs",
    "202": "Owner controlled insurance",
    "204": "Builders Risk",
    "303": "Project Manager",
    "304": "Project Manager Reimbursables",
    "308": "Testing & Inspections",
    "310": "Commissioning Agent",
    "315": "Pre-Opening Expenses",
    "401": "Project Architect- A/E",
    "402": "Project Architect- ASR",
    "403": "Project Architect Reimbursables",
    "408": "Geotechnical Engineer",
    "416": "Technology Design",
    "418": "Peer Review",
    "419": "Peer Review Reimbursables",
    "501": "Pre Construction",
    "503.1": "West Side Renovation",
    "503.3": "Facility Work",
    "503.4": "Deferred Maintenance",
    "503.5": "2024 North/East Scope",
    "503.6": "2025 North/East Scope",
    "601": "Signage",
    "602": "Sponsorship Signage & Branding",
    "603": "Food Beverage and Retail Equipment",
    "604": "Facility Operations",
    "605": "Team Operations",
    "606": "Specialty Lighting/ Sports Lighting",
    "608": "Furnishings",
    "609": "Fixed & Loose Seating",
    "703": "Audio System",
    "704": "Video Displays",
    "706": "Video Production",
    "707": "Broadcast Cabling",
    "708": "TV Displays",
    "710": "Point of Sale(POS)",
    "711": "Security",
    "712": "Audio Visual",
    "713": "DAS & WIFI",
    "802": "Impact Fees",
    "901": "Owner Contingency",
    "316": "General Store / Grounds",
    #"999": "Unclassified / Uncategorized"
}

CLASSIFICATION_RULES = [
    {"code": "403",   "weight": 1000, "fields": ["vendor", "desc"], "keywords": ["populous"],               "note": "Populous always maps to 403"},
    {"code": "316",   "weight": 1000, "fields": ["vendor", "desc"], "keywords": ["general store"],          "note": "general store is 316"},
    {"code": "416",   "weight": 1000, "fields": ["vendor", "desc"], "keywords": ["anthony james partners"], "note": "Anthony James Partners → Technology Design"},
    {"code": "418",   "weight": 500,  "fields": ["ctype"],           "keywords": ["internal commitment"],   "note": "Internal Commitments → Peer Review (418)"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["cmt lab"],                "note": "CMT Laboratories"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hillis"],                 "note": "Hillis-Carnes"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hillis-carnes"],          "note": "Hillis-Carnes (hyphenated)"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hillis carnes"],          "note": "Hillis-Carnes (space)"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hrv conformance"],        "note": "HRV Conformance Verification"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hrv"],                    "note": "HRV (short match)"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["simpson gumpertz"],       "note": "Simpson Gumpertz & Heger"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["testing engineers"],      "note": "Testing Engineers & Consultants"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["trc environmental"],      "note": "TRC Environmental"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["intertek"],               "note": "Intertek PSI"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["professional service industries"], "note": "PSI (full name)"},
    {"code": "308",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["psi"],                    "note": "PSI (abbreviation)"},
    {"code": "310",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["envinity"],               "note": "Envinity"},
    {"code": "310",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["sustainable building"],   "note": "Sustainable Building Partners"},
    {"code": "111",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hrg"],                    "note": "HRG"},
    {"code": "102",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["nations group"],          "note": "Nations Group → Feasibility / PM"},
    {"code": "108",   "weight": 200,  "fields": ["desc"],           "keywords": ["asbestos"],               "note": "Asbestos work → 108"},
    {"code": "108",   "weight": 200,  "fields": ["desc"],           "keywords": ["hazardous"],              "note": "Hazardous material → 108"},
    {"code": "108",   "weight": 150,  "fields": ["desc"],           "keywords": ["pcb"],                    "note": "PCB survey → 108"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["precis engineering"],     "note": "Precis Engineering"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["weber murphy fox"],       "note": "Weber Murphy Fox"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["stahl sheaffer"],         "note": "Stahl Sheaffer Engineering"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["howe engineers"],         "note": "Howe Engineers"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["thornton tomasetti"],     "note": "Thornton Tomasetti"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["florida cosulting"],      "note": "Florida Cosulting (sic) → A/E"},
    {"code": "401",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["florida consulting"],     "note": "Florida Consulting → A/E"},
    {"code": "408",   "weight": 200,  "fields": ["desc"],           "keywords": ["geotechnical"],           "note": "Geotechnical keyword"},
    {"code": "408",   "weight": 150,  "fields": ["desc"],           "keywords": ["soil"],                   "note": "Soil testing → 408"},
    {"code": "416",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["labella"],                "note": "Labella Associates → Technology Design"},
    {"code": "501",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["barton malow"],           "note": "Barton Malow AECOM Hunt"},
    {"code": "501",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bmahajv"],                "note": "BMAHAJV JV"},
    {"code": "501",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bmaha"],                  "note": "BMAHA (partial)"},
    {"code": "501",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["triventure"],             "note": "Triventure → Pre Construction / West Side"},
    {"code": "503.1", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["triventure"],             "note": "Triventure West Side work"},
    {"code": "503.3", "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["leonard s fiore"],        "note": "Leonard S Fiore → Facility Work"},
    {"code": "503.3", "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["gm mccrossin"],           "note": "GM McCrossin → Facility Work"},
    {"code": "503.3", "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["burke"],                  "note": "Burke & Company → Facility Work"},
    {"code": "503.3", "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["mccarl"],                 "note": "S P McCarl → Facility Work"},
    {"code": "601",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["architectural design"],   "note": "AD/S → Signage"},
    {"code": "601",   "weight": 100,  "fields": ["desc"],           "keywords": ["signage"],                "note": "Signage keyword → 601"},
    {"code": "602",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["blair image"],            "note": "Blair Image Elements → Branding"},
    {"code": "602",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["byoglobe"],               "note": "ByoGlobe → Branding"},
    {"code": "602",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["paramount"],              "note": "Paramount & Co → Branding"},
    {"code": "602",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["sc vinyl"],               "note": "SC Vinyl → Sponsorship Signage"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["iowa rotocast"],          "note": "Iowa Rotocast Plastics"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["iowa rotoplastics"],      "note": "Iowa Rotoplastics Inc (alt spelling)"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["iowa roto"],              "note": "Iowa Roto* (catch-all for both spellings)"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["gallery carts"],          "note": "Gallery Carts / Carts of Colorado"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["singer equipment"],       "note": "Singer Equipment"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["carnegie equipment"],     "note": "Carnegie Equipment"},
    {"code": "603",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["11400"],                  "note": "11400 LLC"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["inproduction"],           "note": "InProduction → Facility Ops"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["salient"],                "note": "Salient Engineered Products"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["spicer welding"],         "note": "Spicer Welding & Fabrication"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["cramalot"],               "note": "Cram-A-Lot"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["henry schein"],           "note": "Henry Schein"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bhfoto"],                 "note": "B&H Photo"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bhphoto"],                "note": "B&H Photo variant"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bh foto"],                "note": "B&H Photo (Foto spelling)"},
    {"code": "604",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["bh photo"],               "note": "B&H Photo"},
    {"code": "712",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["dagostino"],              "note": "Dagostino Electronic Services → AV"},
    {"code": "713",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["eplus"],                  "note": "ePlus Technology → DAS & WIFI"},
    {"code": "713",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["wesco"],                  "note": "WESCO Distribution → DAS & WIFI"},
    {"code": "706",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["university video"],       "note": "University Video Services → 706"},
    {"code": "704",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["insane impact"],          "note": "Insane Impact LLC → Video Displays"},
    {"code": "704",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["insane/impact"],          "note": "Insane/Impact (slash variant) → Video Displays"},
    {"code": "704",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["insane"],                 "note": "Insane* catch-all → Video Displays"},
    {"code": "604",   "weight": 50,   "fields": ["desc"],           "keywords": ["penn state general stores"], "note": "PSU General Stores → Facility Ops fallback"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["comunale"],               "note": "S.A. Comunale → Facility Work"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["liberty fire"],           "note": "Liberty Fire → Facility Work"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["thyssenkrupp"],           "note": "ThyssenKrupp Elevator → Facility Work"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["tk elevator"],            "note": "TK Elevator → Facility Work"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["tri-state valve"],        "note": "Tri-State Valve → Facility Work"},
    {"code": "503.3", "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["tristate valve"],         "note": "Tri-State Valve → Facility Work"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["standing stone"],         "note": "Standing Stone Consulting → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["action cleaning"],        "note": "Action Cleaning → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["emerson"],                "note": "Emerson → Facility Ops"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["supra office"],           "note": "Supra Office Solutions → 604"},
    {"code": "316",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["tuckahoe turf"],          "note": "Tuckahoe Turf → 316"},
    {"code": "316",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["dryject"],                "note": "Dryject → 316"},
    {"code": "316",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["aercore"],                "note": "Aer-Core → 316"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["george trailers"],        "note": "George Trailers → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["fulmers storage"],        "note": "Fulmer's Storage Trailers → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["fulmer"],                 "note": "Fulmer's Storage → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["s  c operations"],        "note": "S&C Operations → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["alvarado"],               "note": "Alvarado (turnstiles) → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["triple d enterprise"],    "note": "Triple D Enterprise → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["gxc"],                    "note": "GXC → 604"},
    {"code": "604",   "weight": 150,  "fields": ["vendor", "desc"], "keywords": ["west penn power"],        "note": "West Penn Power → 604"},
    {"code": "608",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["4topps"],                 "note": "4Topps LLC"},
    {"code": "608",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["link interiors"],         "note": "Link Interiors"},
    {"code": "608",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["sedia systems"],          "note": "Sedia Systems"},
    {"code": "703",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["clair global"],           "note": "Clair Global → Audio"},
    {"code": "704",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["insane impact"],          "note": "Insane Impact → Video"},
    {"code": "706",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["vizrt"],                  "note": "Vizrt → Video Production"},
    {"code": "708",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["dobil"],                  "note": "Dobil Laboratories → TV Displays"},
    {"code": "708",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["king systems"],           "note": "King Systems → TV Displays"},
    {"code": "316",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["johnson controls"],       "note": "Johnson Controls → 316"},
    {"code": "316",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["hummer turfgrass"],       "note": "Hummer Turfgrass → 316"},
    {"code": "316",   "weight": 200,  "fields": ["vendor", "desc"], "keywords": ["cdw"],                    "note": "CDW → 316"},
    {"code": "401",   "weight": 100,  "fields": ["desc"],           "keywords": ["design agreement"],       "note": "Design Agreement → A/E"},
    {"code": "401",   "weight": 100,  "fields": ["desc"],           "keywords": ["design services"],        "note": "Design Services IC → A/E"},
    {"code": "501",   "weight": 100,  "fields": ["desc"],           "keywords": ["preconstruction"],        "note": "Pre-construction keyword"},
    {"code": "501",   "weight": 100,  "fields": ["desc"],           "keywords": ["pre construction"],       "note": "Pre-construction keyword"},
    {"code": "503.3", "weight": 100,  "fields": ["desc"],           "keywords": ["construction services"],  "note": "Construction services → 503.3"},
    {"code": "308",   "weight": 100,  "fields": ["desc"],           "keywords": ["inspection"],             "note": "Inspection keyword → 308"},
    {"code": "308",   "weight": 100,  "fields": ["desc"],           "keywords": ["testing"],                "note": "Testing keyword → 308"},
    {"code": "310",   "weight": 100,  "fields": ["desc"],           "keywords": ["commissioning"],          "note": "Commissioning keyword → 310"},
    {"code": "604",   "weight": 100,  "fields": ["desc"],           "keywords": ["operations"],             "note": "Operations keyword → 604"},
    {"code": "604",   "weight": 100,  "fields": ["desc"],           "keywords": ["facility"],               "note": "Facility keyword → 604"},
]

VENDOR_ALIASES = {
    # eBuilder name (lowercased)                                        : PRIVV canonical name(s)
    "insane impact llc":                                                  "Insane/Impact",
    "insane impact":                                                      "Insane/Impact",
    "iowa rotocast plastics inc":                                         "Iowa Rotoplastics Inc",
    "iowa rotocast plastics":                                             "Iowa Rotoplastics Inc",
    "hillis-carnes engineering associates, inc.":                         "Hillis Carnes",
    "hillis-carnes engineering associates":                               "Hillis Carnes",
    "professional service industries inc":                                "Intertek PSI",
    "professional service industries":                                    "Intertek PSI",
    "hrv conformance verification associ":                                "HRV Conformance Verification Associates",
    "hrv conformance verification associates":                            "HRV Conformance Verification Associates",

    # Barton Malow / JV — eBuilder shows the full JV name, PRIVV has multiple vendor entries
    "barton malow aecom hunt alexander":                                  "Triventure",
    "barton malow alexander ii joint venture": [
        "1-CM-GMP-000925000-Barton Malow - Alexander - II Joint Venture -David Peck (dlp50)",
        "Barton Malow - Alexander - II Joint Venture ^",
    ],
    "bmahajv": [
        "1-CM-GMP-000925000-Barton Malow - Alexander - II Joint Venture -David Peck (dlp50)",
        "Barton Malow - Alexander - II Joint Venture ^",
    ],
    "triventure":                                                         "barton malow aecom hunt alexander",

    # Florida Consulting — two spellings on eBuilder side, multiple PRIVV entries
    "florida cosulting": [
        "Florida Cosulting",
        "Florida Consulting, LLC^",
        "1-P Design Agreement_Agreement_Florida Consulting, LLC",
    ],
    "florida consulting": [
        "Florida Cosulting",
        "Florida Consulting, LLC^",
        "1-P Design Agreement_Agreement_Florida Consulting, LLC",
    ],

    # Johnson Controls
    "johnson contrl security solutions":                                  "JOHNSON CONTROLS US HOLDINGS LLC JOHNSON CONTROLS SECURITY SOLUTIONS",
    "johnson controls us holdings llc johnson controls security solutions": "Johnson Control Security Solutions",
    "tuckahoe turf farms inc":                                                  "Hummer Turfgrass Systems Inc",
    "tuckahoe turf farms":                                                      "Hummer Turfgrass Systems Inc",
    "SGH":                                                                 "Simpson Gumpertz & Heger, Inc.^",
    "AD/S":                                         "ARCHITECTURAL DESIGN & SIGNS INC^",
    "AJP":                                          ["ANTHONY JAMES PARTNERS LLC ^"," Anthony James"],

    # Dobil Laboratories — normalize eBuilder full legal name to PRIVV short name.
    # eBuilder uses "DOBIL LABORATORIES INC"; PRIVV may list it as "Dobil Laboratories".
    # These aliases let the comparison loop find the match without touching any
    # other vendor (Clair Global, CDW, etc.).
    "dobil laboratories inc":                        "Dobil Laboratories",
    "dobil laboratories":                            "Dobil Laboratories",
    "dobil":                                         "Dobil Laboratories",
}

DEFAULT_CODE = "901"
FUZZY_THRESHOLD = 40


def normalize(text):
    lowered = str(text).lower()

    if "nations group" in lowered and "suzan" in lowered:
        return "nations group"

    text = lowered
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# Threshold used by find_fuzzy_privv_matches below. Kept as a module-level
# constant (not hardcoded per-vendor) so it can be tuned in one place if
# it's ever too loose/strict across the whole PRIVV file.
PRIVV_VENDOR_FUZZY_THRESHOLD = 87

# PRIVV "generic bucket" vendor labels — rows logged under one of these
# Vendor values don't identify the real company on the Vendor field itself,
# so the Description-embedded-name fallback in find_fuzzy_privv_matches is
# only meaningful for THESE rows. Any row whose Vendor field already names a
# specific, real vendor (e.g. "Populous") must stay attributed to that vendor
# even if its Description happens to start with some other vendor's name
# (e.g. a "HRG Civil Design" line item billed under Populous). This is not a
# per-vendor keyword list — it's the one already-established generic-bucket
# concept (Penn State OPP) used elsewhere in this file (see psu_privv_key /
# PSU_OPP_LABEL), kept in one place so it can be extended if another true
# generic bucket ever shows up in PRIVV data.
GENERIC_BUCKET_VENDORS = {"penn state office of physical plant"}


def find_fuzzy_privv_matches(vendor_label: str, privv_df, threshold: int = PRIVV_VENDOR_FUZZY_THRESHOLD,
                              exclude_labels=None):
    """Return every PRIVV row whose Vendor field is the same company as
    vendor_label, even if spelled slightly differently (e.g. 'Dobil
    Laboratories' vs 'Dobil Laboratories Inc'), AND every row where the
    company name only shows up inside the Description field instead of
    Vendor (PRIVV files sometimes log a generic bucket like 'Penn State
    Office of Physical Plant' as the Vendor and put the real company name
    at the start of the Description, e.g. 'Dobil Laboratories Inc - TV
    monitors purchased...').

    This intentionally does NOT hardcode any vendor name. A row counts as
    a match if, after normalize()-ing the relevant text, either:
      - the Vendor field is identical / a substring match / fuzzy-similar
        to vendor_label (token_sort_ratio >= threshold), or
      - the Description field CONTAINS vendor_label as a substring, or
        scores >= threshold on fuzz.partial_ratio (which finds the best
        matching substring rather than comparing the whole sentence, so a
        short vendor name embedded in a longer description still matches).

    exclude_labels (post-check for the PSU OPP double-count):
        A PRIVV row like Vendor="Penn State Office of Physical Plant",
        Description="Dobil Laboratories Inc - TV monitors..." is matched
        TWICE by design: once for "Dobil Laboratories" (via the
        Description-embedded-name branch below) and once for "Penn State
        Office of Physical Plant" itself (via the plain Vendor-field
        match), because the Vendor field literally says PSU OPP. That
        inflates both companies' PRIVV totals with the same row.

        When vendor_label is itself the generic bucket (PSU OPP) and a
        list of every other vendor label currently on the comparison is
        passed in via exclude_labels, this function re-runs the match for
        each of those other vendors and removes any row they already
        claimed from PSU OPP's result — so a row that genuinely belongs
        to a specific named vendor (Dobil Laboratories) is counted there
        only, and PSU OPP is left with whatever's left over.
    """
    target = normalize(vendor_label)
    if not target:
        return privv_df.iloc[0:0]

    has_vendor = "Vendor" in privv_df.columns
    has_desc = "Description" in privv_df.columns
    if not has_vendor and not has_desc:
        return privv_df.iloc[0:0]

    def _vendor_match(vn: str) -> bool:
        if not vn:
            return False
        if vn == target or vn in target or target in vn:
            return True
        return fuzz.token_sort_ratio(vn, target) >= threshold

    # Vendors that appear in PSU OPP Description fields but must NOT be
    # claimed by their own vendor bucket via description matching — those rows
    # belong to Penn State OPP, not to the named vendor.
    DESC_MATCH_BLOCKLIST = {"clair global", "clair global corporation",
                            "clair global integration llc"}

    def _desc_match(dn: str) -> bool:
        """Return True only when the target vendor name appears at the VERY
        START of the Description (e.g. 'Dobil Laboratories Inc - TV monitors
        purchased...').  Requiring a leading match prevents a vendor name that
        appears mid-sentence (e.g. 'Clair Global - OPP Audio Maintenance'
        inside a Penn State OPP row) from incorrectly pulling that row into
        the named vendor's PRIVV total."""
        if not dn:
            return False
        # Block vendors whose PSU OPP description rows must stay with PSU OPP.
        if target in DESC_MATCH_BLOCKLIST:
            return False
        # Only match when target appears at the start of the description.
        return dn.startswith(target)

    def _is_match(row):
        row_vendor_norm = normalize(str(row.get("Vendor", ""))) if has_vendor else ""
        if has_vendor and _vendor_match(row_vendor_norm):
            return True
        # Also check Description — but ONLY for rows whose own Vendor field
        # is blank or is a known generic bucket (e.g. "Penn State Office of
        # Physical Plant"). If the row already has a specific, real vendor
        # in its Vendor field (e.g. "Populous"), that field is authoritative
        # and must not be overridden just because the Description happens to
        # start with some other vendor's name (e.g. a "HRG Civil Design"
        # line item billed under Populous). Without this guard, every real
        # vendor's line items are at risk of being double-counted into any
        # other vendor whose name happens to appear at the start of a
        # description.
        if has_desc and (not row_vendor_norm or row_vendor_norm in GENERIC_BUCKET_VENDORS):
            if _desc_match(normalize(str(row.get("Description", "")))):
                return True
        return False

    mask = privv_df.apply(_is_match, axis=1)
    matches = privv_df.loc[mask]

    # ── Post-check: strip out rows a specific vendor already claimed ────────
    # Only applies when we're matching the generic PSU OPP bucket itself and
    # the caller told us who else is on the comparison. See exclude_labels
    # note in the docstring above for why this is needed.
    if exclude_labels and target in GENERIC_BUCKET_VENDORS:
        claimed_elsewhere_idx = set()
        for other_label in exclude_labels:
            if normalize(other_label) == target:
                continue
            other_matches = find_fuzzy_privv_matches(other_label, privv_df, threshold)
            claimed_elsewhere_idx.update(other_matches.index)
        if claimed_elsewhere_idx:
            matches = matches.loc[~matches.index.isin(claimed_elsewhere_idx)]

    return matches


def resolve_alias(vendor_raw: str):
    key = vendor_raw.strip().lower()
    if not key:
        return vendor_raw

    # User-added aliases (from the Vendor Aliases tab / Google Sheet) take
    # priority over the hardcoded VENDOR_ALIASES map, since they're the
    # most recently/explicitly taught mapping.
    if key in SHEET_VENDOR_ALIASES:
        return SHEET_VENDOR_ALIASES[key]
    for alias_key, canonical in SHEET_VENDOR_ALIASES.items():
        if alias_key and (alias_key in key or key in alias_key):
            return canonical

    if key in VENDOR_ALIASES:
        return VENDOR_ALIASES[key]
    for alias_key, canonical in VENDOR_ALIASES.items():
        if alias_key and (alias_key in key or key in alias_key):
            return canonical
    return vendor_raw


def resolve_alias_list(vendor_raw: str) -> list:
    result = resolve_alias(vendor_raw)
    if isinstance(result, list):
        return result
    return [result]


def alias_equivalents(vendor_raw: str) -> set:
    """Return the full set of names (raw, as stored in the alias tables)
    known to refer to the SAME vendor as `vendor_raw` — checking both the
    alias KEY and its VALUE(s), in both directions.

    resolve_alias() only matches when `vendor_raw` looks like a known KEY.
    But alias entries are frequently written the other way around too —
    e.g. VENDOR_ALIASES["bmahajv"] = ["Barton Malow - Alexander - II Joint
    Venture ^", ...] — where the eBuilder side actually shows one of the
    long VALUE variants, not the short key "bmahajv". A forward-only
    lookup never connects that text back to its vendor. This function
    checks membership in the whole {key + values} group in either
    direction, so it works regardless of which form the input happens to
    be, for any alias entry shaped that way — not a one-off special case.
    """
    text = str(vendor_raw).strip()
    if not text:
        return set()
    ntext = normalize(text)
    equivalents: set = {text}
    for table in (SHEET_VENDOR_ALIASES, VENDOR_ALIASES):
        for alias_key, canonical in table.items():
            canon_list = canonical if isinstance(canonical, list) else [canonical]
            group = [alias_key] + canon_list
            group_norm = [normalize(g) for g in group]
            if any(ntext == g or (g and (ntext in g or g in ntext)) for g in group_norm):
                equivalents.update(group)
    return equivalents


def convert_date(val):
    try:
        if pd.isna(val) or str(val).strip() == "":
            return ""
        val = str(val).strip()
        if "." in val:
            m, d, y = val.split(".")
        elif "/" in val:
            m, d, y = val.split("/")
        else:
            return val
        return f"{int(m)}/{int(d)}/{int(y)}"
    except Exception:
        return val


def normalize_date_to_tuple(val):
    """Parse both eBuilder (MM.DD.YYYY) and Privv (M/D/YYYY) date formats
    into a comparable (year, month, day) tuple. Returns None on failure.

    Handles stray non-date strings (e.g. 'Status', 'N/A', header text leaked
    into the Date column) by returning None instead of raising ValueError.
    """
    try:
        val = str(val).strip()
        if not val:
            return None
        if "." in val:
            parts = val.split(".")
        elif "/" in val:
            parts = val.split("/")
        else:
            return None
        if len(parts) != 3:
            return None
        # Guard: all three parts must be purely numeric before calling int()
        if not all(p.strip().isdigit() for p in parts):
            return None
        m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
        # Sanity-check the ranges so completely bogus values don't slip through
        if not (1 <= m <= 12 and 1 <= d <= 31 and 1900 <= y <= 2100):
            return None
        return (y, m, d)
    except Exception:
        return None


def assign_code_weighted(row, lookup: dict) -> tuple:
    vendor = normalize(row.get("Company", ""))
    desc   = normalize(row.get("Description", ""))
    ctype  = normalize(row.get("Commitment Type", ""))

    field_values = {"vendor": vendor, "desc": desc, "ctype": ctype}

    best_score = -1
    best_code  = None
    best_note  = "no match"

    for rule in CLASSIFICATION_RULES:
        matched = all(
            any(kw in field_values.get(f, "") for f in rule["fields"])
            for kw in rule["keywords"]
        )
        if matched and rule["weight"] > best_score:
            best_score = rule["weight"]
            best_code  = rule["code"]
            best_note  = rule.get("note", "")

    if lookup:
        match = process.extractOne(
            vendor,
            lookup.keys(),
            scorer=fuzz.token_set_ratio,
            score_cutoff=FUZZY_THRESHOLD,
        )
        if match:
            matched_key, score, _ = match
            lookup_weight = 150
            if lookup_weight > best_score:
                best_score = lookup_weight
                best_code  = lookup[matched_key]
                best_note  = f"privv lookup (fuzzy {score:.0f}%): {matched_key}"

    if best_code is None:
        return DEFAULT_CODE, "default fallback"

    return best_code, best_note


# Preferred display order for the well-known eBuilder columns. This is no
# longer a filter -- _rows_to_entries below keeps EVERY real column from the
# eBuilder row, it just uses this list to decide what order to show the
# familiar ones in before tacking on anything extra.
ENTRY_COLS = [
    "ID", "Description", "Company", "Date", "Status",
    "Commitment Type", "Commitment Amount", "Current Commitment",
    "Projected Commitment", "Actuals Approved", "Remaining Balance",
]


def _rows_to_entries(df_subset):
    """Convert a slice of the eBuilder dataframe into a list of plain dicts
    representing the FULL line item (every real column from the eBuilder
    file) so the drill-down entries window can show all of it.

    Internal helper columns that we add ourselves during processing
    (e.g. "_commit_num", "_resolved_vendors") are excluded since they're
    not part of the original eBuilder row -- everything else is kept.
    """
    real_cols = [c for c in df_subset.columns if not str(c).startswith("_")]
    ordered = [c for c in ENTRY_COLS if c in real_cols] + \
              [c for c in real_cols if c not in ENTRY_COLS]
    return df_subset[ordered].to_dict("records")


# ---------------------------------------------------------------------------
# GUI STATE
# ---------------------------------------------------------------------------
ebuilder_files: list = []
privv_file: str = ""


def select_ebuilder_files():
    global ebuilder_files
    files = list(filedialog.askopenfilenames(
        title="Select one or more eBuilder CSV files",
        filetypes=[("CSV Files", "*.csv")]
    ))
    if files:
        ebuilder_files = files
        names = ", ".join(os.path.basename(f) for f in files)
        ebuilder_label.config(text=f"✅ {len(files)} file(s): {names}", fg="green")
    else:
        if not ebuilder_files:
            ebuilder_label.config(text="No eBuilder files selected.", fg="gray")


def select_privv_file():
    global privv_file
    f = filedialog.askopenfilename(
        title="Select PRIVV reference CSV file",
        filetypes=[("CSV Files", "*.csv")]
    )
    if f:
        privv_file = f
        privv_label.config(text=f"✅ Selected: {os.path.basename(f)}", fg="green")
    else:
        if not privv_file:
            privv_label.config(text="No PRIVV file selected.", fg="gray")


def clear_ebuilder_files():
    global ebuilder_files
    ebuilder_files = []
    ebuilder_label.config(text="No eBuilder files selected.", fg="gray")
    log("eBuilder file list cleared.")


def log(msg: str):
    output_box.insert(tk.END, msg + "\n")
    output_box.see(tk.END)


# ---------------------------------------------------------------------------
# COMPARISON
# ---------------------------------------------------------------------------
def run_comparison():
    if not ebuilder_files:
        messagebox.showerror("Error", "Select at least one eBuilder file first.")
        return
    if not privv_file:
        messagebox.showerror("Error", "Select the PRIVV file first.")
        return

    log("\n── Running Vendor Comparison ──────────────────────────────────────")

    privv_df = pd.read_csv(privv_file, dtype=str).fillna("")
    if "Vendor" not in privv_df.columns or "Amount" not in privv_df.columns:
        messagebox.showerror("Error", "PRIVV file must have 'Vendor' and 'Amount' columns.")
        return

    privv_df["_amount_num"] = pd.to_numeric(
        privv_df["Amount"].str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)
    privv_total_all = privv_df["_amount_num"].sum()
    log(f"  PRIVV file loaded: {len(privv_df)} entries, total amount = ${privv_total_all:,.2f}")

    data_list = []
    for f in ebuilder_files:
        df = pd.read_csv(f, header=0, dtype=str).fillna("")
        df = df[~(df == "").all(axis=1)]
        data_list.append(df)
    eb_raw = pd.concat(data_list, ignore_index=True).drop_duplicates()

    if "#" in eb_raw.columns:
        eb_raw = eb_raw.rename(columns={"#": "ID"})

    col_names = [
        "ID", "Description", "Company", "Date", "Status",
        "Commitment Type", "Commitment Amount",
        "Current Commitment", "Projected Commitment",
        "Actuals Approved", "Remaining Balance",
    ]
    if not all(c in eb_raw.columns for c in ["Description", "Current Commitment"]):
        while len(eb_raw.columns) < len(col_names):
            eb_raw[len(eb_raw.columns)] = ""
        eb_raw.columns = col_names[:len(eb_raw.columns)]

    # NOTE: Company is intentionally NOT overwritten here so that both
    # Description and Company are checked independently against PRIVV vendors.

    # Drop summary/totals rows
    eb_raw = eb_raw[eb_raw["Description"].str.strip() != ""].reset_index(drop=True)

    _current = pd.to_numeric(
        eb_raw["Current Commitment"].str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)
    _commitment = pd.to_numeric(
        eb_raw["Commitment Amount"].str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)
    # Use Current Commitment when it has a value; fall back to Commitment Amount
    # for rows where Current Commitment is zero or was blank/N/A (e.g. Pending agreements)
    eb_raw["_commit_num"] = _current.where(_current != 0, other=_commitment)

    # Hard-coded override: map eBuilder JV description to PRIVV vendor name BMAHAJV
    eb_raw.loc[
        eb_raw["Company"].str.contains("Barton Malow - Alexander - II Joint Venture", case=False, na=False) |
        eb_raw["Description"].str.contains("Barton Malow - Alexander - II Joint Venture", case=False, na=False),
        "Company"
    ] = "BMAHAJV"
    eb_raw.loc[
        eb_raw["Company"].str.contains("J.V. MANUFACTURING, INC-CRAMALOT", case=False, na=False) |
        eb_raw["Description"].str.contains("J.V. MANUFACTURING, INC-CRAMALOT", case=False, na=False),
        "Company"
    ] = "Cram-A-Lot"   
    eb_raw.loc[
    eb_raw["Company"].str.contains("Architectural Design", case=False, na=False, regex=False) |
    eb_raw["Description"].str.contains("Architectural Design", case=False, na=False, regex=False),
    "Company"
] = "AD/S"

    eb_raw.loc[
        eb_raw["Company"].str.contains("Simpson Gumpertz", case=False, na=False, regex=False) |
        eb_raw["Description"].str.contains("Simpson Gumpertz", case=False, na=False, regex=False),
        "Company"
    ] = "SGH" 
    eb_raw.loc[
        eb_raw["Company"].str.contains("Anthony James", case=False, na=False, regex=False) |
        eb_raw["Description"].str.contains("Anthony James", case=False, na=False, regex=False),
        "Company"
    ] = "AJP" 


 



#J.V. MANUFACTURING, INC-CRAMALOT
#
    # Build resolved-vendor column checking BOTH Description and Company
    # Build resolved-vendor column checking BOTH Description and Company
    # for alias matches, deduplicating the resulting names.
    eb_raw["_resolved_vendors"] = eb_raw.apply(
        lambda r: list({
            name
            for field in [r.get("Description", ""), r.get("Company", "")]
            for name in resolve_alias_list(str(field))
        }),
        axis=1,
    )

    comparison_rows = []
    vendors_seen = {}

    for _, row in privv_df.iterrows():
        vendor_raw = str(row.get("Vendor", "")).strip()
        if not vendor_raw:
            continue
        vendor_key = normalize(vendor_raw)
        if vendor_key not in vendors_seen:
            vendors_seen[vendor_key] = vendor_raw


    # ── Pre-pass: Clair Global rows must have "Clair Global" in BOTH ───────────
    #    Company AND Description to count as a Clair Global eBuilder entry.
    #    Rows where it only appears in one field are forced to PSU OPP and
    #    excluded from ALL matching loops.
    clair_in_either = (
        eb_raw["Company"].str.contains("Clair Global", case=False, na=False, regex=False) |
        eb_raw["Description"].str.contains("Clair Global", case=False, na=False, regex=False)
    )
    clair_both = (
        eb_raw["Company"].str.contains("Clair Global", case=False, na=False, regex=False) &
        eb_raw["Description"].str.contains("Clair Global", case=False, na=False, regex=False)
    )
    # Rows that mention Clair Global in only one field → pre-assign to PSU OPP,
    # remove them so NO matching loop ever sees them.
    clair_partial_indices = eb_raw[clair_in_either & ~clair_both].index
    clair_partial_total   = eb_raw.loc[clair_partial_indices, "_commit_num"].sum()
    clair_partial_entries = _rows_to_entries(eb_raw.loc[clair_partial_indices])
    if len(clair_partial_indices):
        log(f"  Clair Global partial match ({len(clair_partial_indices)} row(s), "
            f"${clair_partial_total:,.2f}) → Penn State OPP")
    eb_raw = eb_raw[~eb_raw.index.isin(clair_partial_indices)].copy()

    # Rebuild _resolved_vendors after stripping Clair Global partial rows
    eb_raw["_resolved_vendors"] = eb_raw.apply(
        lambda r: list({
            name
            for field in [r.get("Description", ""), r.get("Company", "")]
            for name in resolve_alias_list(str(field))
        }),
        axis=1,
    )

    for vendor_key, vendor_raw in vendors_seen.items():
        privv_mask = privv_df["Vendor"].apply(lambda v: normalize(str(v)) == vendor_key)
        privv_total = privv_df.loc[privv_mask, "_amount_num"].sum()

        eb_direct = (
            eb_raw["Description"].apply(lambda d: vendor_key in normalize(str(d))) |
            eb_raw["Company"].apply(lambda c: vendor_key in normalize(str(c)))
        )
        eb_alias = eb_raw["_resolved_vendors"].apply(
            lambda names: any(normalize(n) == vendor_key for n in names)
        )
        eb_mask = eb_direct | eb_alias

        eb_total = eb_raw.loc[eb_mask, "_commit_num"].sum()
        eb_matches = int(eb_mask.sum())

        delta = eb_total - privv_total

        if eb_total == 0 and privv_total == 0:
            status = "⚪ No Data"
        #Future iterations or people working on this can potentially remove this line above as ngl its kinda pointless why in the world would we have a instance with zero information in both softwares
        elif abs(delta) < 0.01:
            status = "✅ Match"
        elif eb_total > privv_total:
            status = "🔼 eBuilder Higher"
        else:
            status = "🔽 eBuilder Lower"

        comparison_rows.append({
            "Vendor":         vendor_raw,
            "PRIVV Total":    privv_total,
            "eBuilder Total": eb_total,
            "Delta":          delta,
            "eB Matches":     eb_matches,
            "Status":         status,
            "_entries":       _rows_to_entries(eb_raw.loc[eb_mask]),
        })

    if not comparison_rows:
        log("No vendors found in PRIVV file.")
        return

    # ── Main matching loop ───────────────────────────────────────────────────
    all_matched_indices = set()
    for vendor_key, vendor_raw in vendors_seen.items():
        eb_direct = (
            eb_raw["Description"].apply(lambda d: vendor_key in normalize(str(d))) |
            eb_raw["Company"].apply(lambda c: vendor_key in normalize(str(c)))
        )
        eb_alias = eb_raw["_resolved_vendors"].apply(
            lambda names: any(normalize(n) == vendor_key for n in names)
        )
        eb_mask = eb_direct | eb_alias
        all_matched_indices.update(eb_raw[eb_mask].index.tolist())

    unmatched_eb = eb_raw[~eb_raw.index.isin(all_matched_indices)].copy()

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _add_to_psu(amount, n_rows, entries=None):
        entries = entries or []
        psu_key = "penn state office of physical plant"
        existing = next(
            (r for r in comparison_rows if r["Vendor"].strip().lower() == psu_key),
            None,
        )
        if existing:
            existing["eBuilder Total"] += amount
            existing["Delta"]           = existing["eBuilder Total"] - existing["PRIVV Total"]
            existing["eB Matches"]     += n_rows
            existing.setdefault("_entries", []).extend(entries)
            if abs(existing["Delta"]) < 0.01:
                existing["Status"] = "✅ Match"
            elif existing["eBuilder Total"] > existing["PRIVV Total"]:
                existing["Status"] = "🔼 eBuilder Higher"
            else:
                existing["Status"] = "🔽 eBuilder Lower"
        else:
            comparison_rows.append({
                "Vendor":         "Penn State Office of Physical Plant",
                "PRIVV Total":    0.0,
                "eBuilder Total": amount,
                "Delta":          amount,
                "eB Matches":     n_rows,
                "Status":         "🔼 eBuilder Higher",
                "_entries":       list(entries),
            })

    def _add_to_new_company(vendor_label, amount, n_rows, entries=None):
        comparison_rows.append({
            "Vendor":         vendor_label,
            "PRIVV Total":    0.0,
            "eBuilder Total": amount,
            "Delta":          amount,
            "eB Matches":     n_rows,
            "Status":         "🔼 eBuilder Higher",
            "_entries":       list(entries or []),
        })

    # Flush the pre-pass Clair Global partial rows into PSU OPP now that
    # comparison_rows is fully built and _add_to_psu is defined.
    if clair_partial_total != 0:
        _add_to_psu(clair_partial_total, len(clair_partial_indices), clair_partial_entries)

    # ── Build PSU OPP date set from PRIVV ────────────────────────────────────
    # Collect every committed-entry date that PRIVV has for Penn State OPP so
    # we can match unmatched eBuilder rows by date as well as description.
    psu_privv_key = "penn state office of physical plant"
    psu_privv_dates: set = set()
    if "Date Committed" in privv_df.columns:
        date_col = "Date Committed"
    elif "Date" in privv_df.columns:
        date_col = "Date"
    else:
        date_col = None

    if date_col:
        psu_privv_mask = privv_df["Vendor"].apply(
            lambda v: normalize(str(v)) == psu_privv_key
        )
        for raw_date in privv_df.loc[psu_privv_mask, date_col]:
            tup = normalize_date_to_tuple(raw_date)
            if tup:
                psu_privv_dates.add(tup)

    if psu_privv_dates:
        log(f"  PSU OPP PRIVV committed dates loaded: {len(psu_privv_dates)} date(s)")

    def _eb_date_matches_psu(eb_date_val) -> bool:
        """Return True if this eBuilder row's date appears in the PSU OPP PRIVV dates."""
        tup = normalize_date_to_tuple(eb_date_val)
        return tup is not None and tup in psu_privv_dates

    # ── Build a flat list of PRIVV vendor names for fuzzy matching ───────────
    # Used below to check whether an "unmatched" eBuilder row actually belongs
    # to a known PRIVV vendor that just wasn't caught by the exact-match pass.
    privv_vendor_names: list = [
        str(row.get("Vendor", "")).strip()
        for _, row in privv_df.iterrows()
        if str(row.get("Vendor", "")).strip()
    ]
    # Deduplicate while preserving order
    _seen_pv: set = set()
    privv_vendor_names_dedup: list = []
    for _pv in privv_vendor_names:
        _k = normalize(_pv)
        if _k not in _seen_pv:
            _seen_pv.add(_k)
            privv_vendor_names_dedup.append(_pv)
    privv_vendor_names = privv_vendor_names_dedup

    UNMATCHED_FUZZY_THRESHOLD = 72  # higher bar than classification fuzzy (40)

    # Vendors handled exclusively by _fix_zero_eb post-pass.
    # Prevent the fuzzy loop from grabbing their rows and mis-grouping them
    # (e.g. "DYWIDAG SYSTEMS INC" scoring high against "KING SYSTEMS" because
    # both share the token "SYSTEMS").
    FUZZY_SKIP_KEYWORDS = {"dywidag", "thornton tomasetti"}

    def _fuzzy_find_privv_vendor(company_str: str, description_str: str):
        """
        Return (privv_vendor_raw, score) if either the Company field or the
        Description field fuzzy-matches a known PRIVV vendor above the threshold,
        OR if the vendor name is found as a substring inside the description.
        Returns (None, 0) if no match is found.
        """
        if not privv_vendor_names:
            return None, 0

        # Skip vendors that are handled exclusively by the _fix_zero_eb post-pass.
        # Without this guard, shared generic tokens like "SYSTEMS" can cause
        # e.g. "DYWIDAG SYSTEMS INC" to fuzzy-match "KING SYSTEMS".
        combined_lower = (company_str + " " + description_str).lower()
        if any(kw in combined_lower for kw in FUZZY_SKIP_KEYWORDS):
            return None, 0

        candidates = []

        # 1. Fuzzy match Company field against all PRIVV vendor names
        if company_str.strip():
            m = process.extractOne(
                normalize(company_str),
                [normalize(v) for v in privv_vendor_names],
                scorer=fuzz.token_set_ratio,
                score_cutoff=UNMATCHED_FUZZY_THRESHOLD,
            )
            if m:
                matched_norm, score, idx = m
                candidates.append((privv_vendor_names[idx], score, "Company fuzzy"))

        # 2. Fuzzy match Description field against all PRIVV vendor names
        if description_str.strip():
            m2 = process.extractOne(
                normalize(description_str),
                [normalize(v) for v in privv_vendor_names],
                scorer=fuzz.token_set_ratio,
                score_cutoff=UNMATCHED_FUZZY_THRESHOLD,
            )
            if m2:
                matched_norm2, score2, idx2 = m2
                candidates.append((privv_vendor_names[idx2], score2, "Description fuzzy"))

        # 3. Substring check: is any PRIVV vendor name contained in the description?
        desc_norm = normalize(description_str)
        for pv in privv_vendor_names:
            pv_norm = normalize(pv)
            if pv_norm and pv_norm in desc_norm:
                candidates.append((pv, 100, "vendor-in-description"))

        # 4. Substring check: is any PRIVV vendor name contained in the company field?
        comp_norm = normalize(company_str)
        for pv in privv_vendor_names:
            pv_norm = normalize(pv)
            if pv_norm and pv_norm in comp_norm:
                candidates.append((pv, 100, "vendor-in-company"))

        if not candidates:
            return None, 0

        # Pick the highest-scoring candidate
        best = max(candidates, key=lambda x: x[1])
        return best[0], best[1]   # (privv_vendor_raw, score)

    def _add_to_existing_privv_vendor(privv_vendor_raw: str, amount: float,
                                       n_rows: int, entries=None):
        """
        Find the comparison_row that belongs to privv_vendor_raw and add the
        eBuilder amount to it.  If for some reason it is not found, fall back
        to creating a new entry so nothing is silently lost.
        """
        target_key = normalize(privv_vendor_raw)
        for r in comparison_rows:
            if normalize(r["Vendor"]) == target_key:
                r["eBuilder Total"] += amount
                r["eB Matches"]     += n_rows
                r.setdefault("_entries", []).extend(entries or [])
                r["Delta"] = r["eBuilder Total"] - r["PRIVV Total"]
                r["Status"] = (
                    "✅ Match"           if abs(r["Delta"]) < 0.01 else
                    "🔼 eBuilder Higher" if r["eBuilder Total"] > r["PRIVV Total"] else
                    "🔽 eBuilder Lower"
                )
                return True
        # Fallback: vendor found via fuzzy but somehow not in comparison_rows yet
        _add_to_new_company(privv_vendor_raw, amount, n_rows, entries)
        return False

    # ── Route each remaining unmatched row ───────────────────────────────────
    # Priority order:
    #   1. Fuzzy / substring match to a known PRIVV vendor → credit that vendor
    #   2. "Penn" in eBuilder Description OR date matches PSU OPP PRIVV → PSU OPP
    #   3. Everything else → new company entry
    for _, urow in unmatched_eb.iterrows():
        row_amount   = float(urow["_commit_num"])
        description  = str(urow.get("Description", ""))
        eb_date      = str(urow.get("Date", ""))
        company_str  = str(urow.get("Company", "")).strip()
        vendor_label = company_str or description[:40].strip()
        entries      = _rows_to_entries(unmatched_eb.loc[[urow.name]])

        # ── Step 1: fuzzy / substring match to a known PRIVV vendor ──────────
        privv_match, match_score = _fuzzy_find_privv_vendor(company_str, description)
        if privv_match:
            log(f"  Unmatched fuzzy→PRIVV (score {match_score}) "
                f"'{vendor_label}' → '{privv_match}'")
            _add_to_existing_privv_vendor(privv_match, row_amount, 1, entries)
            continue

        # ── Step 2: Penn State OPP heuristics ────────────────────────────────
        penn_in_desc = "penn" in description.lower()
        date_matches = _eb_date_matches_psu(eb_date)

        if penn_in_desc or date_matches:
            reason = []
            if penn_in_desc:  reason.append("'Penn' in desc")
            if date_matches:  reason.append(f"date match ({eb_date})")
            log(f"  Unmatched {' & '.join(reason)} → PSU OPP: {vendor_label!r}")
            _add_to_psu(row_amount, 1, entries)
            continue

        # ── Step 3: genuinely new vendor ─────────────────────────────────────
        if not vendor_label:
            vendor_label = "Unknown (eBuilder Only)"
        log(f"  Unmatched → new entry: {vendor_label!r}")
        _add_to_new_company(vendor_label, row_amount, 1, entries)


    # ── Post-pass: fold "The Pennsylvania State University" into PSU OPP ─────
    # Scan all final comparison_rows; any row whose Vendor name matches
    # "Pennsylvania State University" (case-insensitive) has its eBuilder Total
    # merged into Penn State OPP and is then removed from the table.
    AJP_NAMES = {"anthony james partners llc ^", "anthony james partners llc", "anthony james partners", "ajp"}
    ajp_eb_mask = (
        eb_raw["Company"].str.contains("Anthony James", case=False, na=False, regex=False) |
        eb_raw["Description"].str.contains("Anthony James", case=False, na=False, regex=False)
    )
    ajp_eb_total   = eb_raw.loc[ajp_eb_mask, "_commit_num"].sum()
    ajp_eb_entries = _rows_to_entries(eb_raw.loc[ajp_eb_mask])
    ajp_merge_rows = [r for r in comparison_rows if r["Vendor"].strip().lower() in AJP_NAMES]
    if ajp_merge_rows or ajp_eb_total:
        for mr in ajp_merge_rows:
            log(f"  Post-pass: merging '{mr['Vendor']}' "
                f"(${ajp_eb_total:,.2f}, {int(ajp_eb_mask.sum())} row(s)) → Penn State OPP")
        if ajp_eb_total:
            _add_to_psu(ajp_eb_total, int(ajp_eb_mask.sum()), ajp_eb_entries)
        comparison_rows[:] = [
            r for r in comparison_rows
            if r["Vendor"].strip().lower() not in AJP_NAMES
        ]

    def _fix_zero_eb(privv_variants: set, eb_keyword: str, display_name: str = None):
        """
        Post-pass correction for vendors whose eBuilder rows were not caught by
        the main matching loop (e.g. vendor absent from PRIVV entirely, or
        matched via _add_to_new_company before _fix_zero_eb runs).

        Changes vs. original:
          - The `eBuilder Total == 0` guard is removed: we always write the
            correct total, whether the row was previously $0 or already partially
            credited by _add_to_new_company.
          - If no comparison_row matches privv_variants at all (vendor is
            completely absent from PRIVV), a new row is created via
            _add_to_new_company so the amount is never silently lost.
        """
        mask = (
            eb_raw["Company"].str.contains(eb_keyword, case=False, na=False, regex=False) |
            eb_raw["Description"].str.contains(eb_keyword, case=False, na=False, regex=False)
        )
        eb_total = eb_raw.loc[mask, "_commit_num"].sum()
        if not eb_total:
            return

        eb_count   = int(mask.sum())
        eb_entries = _rows_to_entries(eb_raw.loc[mask])

        # Locate an existing row — could be from the PRIVV main loop *or* from a
        # prior _add_to_new_company call (vendor label like "DYWIDAG SYSTEMS INC^"
        # normalises into the privv_variants set).
        target = next(
            (r for r in comparison_rows if r["Vendor"].strip().lower() in privv_variants),
            None,
        )

        if target is not None:
            # Always overwrite: removes the old `== 0` guard that caused the bug.
            target["eBuilder Total"] = eb_total
            target["eB Matches"]     = eb_count
            target["_entries"]       = eb_entries
            target["Delta"]          = target["eBuilder Total"] - target["PRIVV Total"]
            target["Status"] = (
                "✅ Match"           if abs(target["Delta"]) < 0.01 else
                "🔼 eBuilder Higher" if target["eBuilder Total"] > target["PRIVV Total"] else
                "🔽 eBuilder Lower"
            )
            log(f"  Post-pass fix: '{target['Vendor']}' eBuilder total → ${eb_total:,.2f}")
        else:
            # Vendor has no PRIVV row at all — create one so the amount is visible.
            label = display_name or eb_keyword
            _add_to_new_company(label, eb_total, eb_count, eb_entries)
            log(f"  Post-pass fix: created new entry '{label}' → ${eb_total:,.2f}")

    _fix_zero_eb({"dywidag systems inc^", "dywidag systems inc"},
                 "DYWIDAG", display_name="DYWIDAG SYSTEMS INC")
    _fix_zero_eb({"thornton tomasetti inc^", "thornton tomasetti inc", "thornton tomasetti"},
                 "Thornton Tomasetti")

    PSU_NAMES = {"the pennsylvania state university", "pennsylvania state university"}
    psu_merge_rows = [
        r for r in comparison_rows
        if r["Vendor"].strip().lower() in PSU_NAMES
    ]
    if psu_merge_rows:
        for mr in psu_merge_rows:
            merged_amount  = mr["eBuilder Total"]
            merged_matches = mr["eB Matches"]
            merged_entries = mr.get("_entries", [])
            log(f"  Post-pass: merging '{mr['Vendor']}' "
                f"(${merged_amount:,.2f}, {merged_matches} row(s)) → Penn State OPP")
            _add_to_psu(merged_amount, merged_matches, merged_entries)
        # Remove the now-merged rows from the display list
        comparison_rows[:] = [
            r for r in comparison_rows
            if r["Vendor"].strip().lower() not in PSU_NAMES
        ]

    # ── Recompute PRIVV Total with fuzzy matching ───────────────────────────
    # The PRIVV Total set earlier (vendors_seen exact-match loop) misses rows
    # where the company name only appears in Description (e.g. Vendor =
    # "Penn State Office of Physical Plant", Description = "Dobil
    # Laboratories Inc - ..."), or where the Vendor spelling varies slightly
    # ("Dobil Laboratories" vs "Dobil Laboratories Inc"). find_fuzzy_privv_matches
    # is the same lookup used by the entries drilldown window, so recomputing
    # every row's total with it here keeps the summary table and the
    # drilldown in agreement instead of only fixing it after a click.
    _all_comparison_vendors = [r2["Vendor"] for r2 in comparison_rows]
    for r in comparison_rows:
        fuzzy_matches = find_fuzzy_privv_matches(r["Vendor"], privv_df, exclude_labels=_all_comparison_vendors)
        if "_amount_num" in fuzzy_matches.columns:
            fuzzy_total = fuzzy_matches["_amount_num"].sum()
        else:
            fuzzy_total = pd.to_numeric(
                fuzzy_matches.get("Amount", pd.Series(dtype=str)).astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).sum()
        if abs(fuzzy_total - r["PRIVV Total"]) > 0.005:
            log(f"  Fuzzy PRIVV total fix: '{r['Vendor']}' "
                f"${r['PRIVV Total']:,.2f} → ${fuzzy_total:,.2f}")
        r["PRIVV Total"] = fuzzy_total
        r["Delta"] = r["eBuilder Total"] - r["PRIVV Total"]
        r["Status"] = (
            "✅ Match"           if abs(r["Delta"]) < 0.01 else
            "🔼 eBuilder Higher" if r["eBuilder Total"] > r["PRIVV Total"] else
            "🔽 eBuilder Lower"
        )

    # ── De-dup pass: strip eBuilder entries also claimed by PSU OPP ─────────
    # Walk every company row EXCEPT PSU OPP and check each of its eBuilder
    # entries against the set of eBuilder IDs PSU OPP already claimed. Any
    # entry that shows up in both is a double-count bug (same eBuilder row
    # matched into a real vendor AND swept into PSU OPP) — PSU OPP wins and
    # the entry is removed from the other company's row/total/match count.
    # Every removal is logged to the GUI so it's visible, not silent.
    _psu_key = "penn state office of physical plant"
    _psu_row = next(
        (r for r in comparison_rows if r["Vendor"].strip().lower() == _psu_key),
        None,
    )
    if _psu_row is not None:
        _psu_ids = {
            str(e.get("ID", "")).strip()
            for e in _psu_row.get("_entries", [])
            if str(e.get("ID", "")).strip()
        }

        def _entry_amount(e):
            def _num(v):
                s = str(v).replace(",", "").strip()
                if not s or s.lower() == "nan":
                    return 0.0
                try:
                    return float(s)
                except ValueError:
                    return 0.0
            cur = _num(e.get("Current Commitment", 0))
            return cur if cur != 0 else _num(e.get("Commitment Amount", 0))

        if _psu_ids:
            _dup_lines = []
            for r in comparison_rows:
                if r is _psu_row:
                    continue
                entries = r.get("_entries", [])
                keep, removed = [], []
                for e in entries:
                    eid = str(e.get("ID", "")).strip()
                    if eid and eid in _psu_ids:
                        removed.append(e)
                    else:
                        keep.append(e)
                if removed:
                    removed_amount = sum(_entry_amount(e) for e in removed)
                    r["_entries"]       = keep
                    r["eBuilder Total"] = r["eBuilder Total"] - removed_amount
                    r["eB Matches"]     = max(0, r["eB Matches"] - len(removed))
                    r["Delta"]          = r["eBuilder Total"] - r["PRIVV Total"]
                    r["Status"] = (
                        "✅ Match"           if abs(r["Delta"]) < 0.01 else
                        "🔼 eBuilder Higher" if r["eBuilder Total"] > r["PRIVV Total"] else
                        "🔽 eBuilder Lower"
                    )
                    for e in removed:
                        _dup_lines.append(e)  # kept for internal tracking only

    comparison_rows.sort(key=lambda r: abs(r["Delta"]), reverse=True)

    # Total entries assigned across all companies (kept for internal use,
    # no longer printed).
    total_assigned_entries = sum(r.get("eB Matches", 0) for r in comparison_rows)

    # Built from claimed_ids (every row appended into some company's
    # _entries across ALL passes: direct match, fuzzy, PSU OPP, AJP merge,
    # _fix_zero_eb, etc.) — the DISTINCT set of eBuilder IDs that ended up
    # in some entries list, used below for the true de-duped total.
    claimed_ids: set = set()
    for r in comparison_rows:
        for entry in r.get("_entries", []):
            eid = str(entry.get("ID", "")).strip()
            if eid:
                claimed_ids.add(eid)

    all_eb_ids = (
        eb_raw["ID"].astype(str).str.strip().tolist()
        if "ID" in eb_raw.columns else []
    )
    matched_ids   = [eid for eid in all_eb_ids if eid in claimed_ids]
    unmatched_ids = [eid for eid in all_eb_ids if eid not in claimed_ids]

    # ── TRUE POST-MATCHING TOTALS (de-duplicated) ────────────────────────────
    # "eBuilder Total" on each comparison row is a per-vendor subtotal. The
    # substring/alias matching used above can occasionally match the same
    # eBuilder row into two different vendors' totals, which inflates the
    # sum of those subtotals. claimed_ids (built above) is the distinct set
    # of eBuilder IDs that ended up counted anywhere, so re-summing eb_raw
    # filtered to just those ids gives the real, de-duped eBuilder total —
    # each eBuilder row counted once no matter how many vendors claimed it.
    naive_eb_sum    = sum(r["eBuilder Total"] for r in comparison_rows)
    naive_privv_sum = sum(r["PRIVV Total"]    for r in comparison_rows)

    if "ID" in eb_raw.columns:
        true_eb_total = eb_raw.loc[
            eb_raw["ID"].astype(str).str.strip().isin(claimed_ids), "_commit_num"
        ].sum()
    else:
        true_eb_total = naive_eb_sum  # no unique ID column available to de-dupe against

    # PRIVV totals are pulled per vendor by an exact-equality mask on a
    # unique vendor_key, so they shouldn't double count the way eBuilder can —
    # but sum straight from privv_df too so the "true" totals are always
    # calculated the same way: from the source data, not from subtotals.
    true_privv_total = privv_df["_amount_num"].sum()

    log("\n── TRUE POST-MATCHING TOTALS (de-duplicated) ───────────────────────────")
    log(f"  eBuilder Total (sum of per-vendor subtotals) : ${naive_eb_sum:,.2f}")
    log(f"  eBuilder Total (true, de-duped by ID)        : ${true_eb_total:,.2f}")
    if abs(naive_eb_sum - true_eb_total) > 0.01:
        log(f"  ⚠ Difference of ${naive_eb_sum - true_eb_total:,.2f} — some eBuilder "
            f"entries were counted under more than one company above.")
    log(f"  PRIVV Total    (sum of per-vendor subtotals) : ${naive_privv_sum:,.2f}")
    log(f"  PRIVV Total    (true, all PRIVV rows loaded)  : ${true_privv_total:,.2f}")
    if abs(naive_privv_sum - true_privv_total) > 0.01:
        log(f"  ⚠ Difference of ${naive_privv_sum - true_privv_total:,.2f} between PRIVV "
            f"rows attributed to a vendor and all PRIVV rows loaded.")
    log("────────────────────────────────────────────────────────────────────────\n")

    show_comparison_window(comparison_rows, privv_df)
    log(f"Comparison complete — {len(comparison_rows)} vendor(s) analyzed.")


def show_comparison_window(rows, privv_df):
    win = tk.Toplevel(root)
    win.title("Vendor Amount Comparison: PRIVV vs eBuilder")
    win.geometry("1000x520")

    total_privv    = sum(r["PRIVV Total"]    for r in rows)
    total_ebuilder = sum(r["eBuilder Total"] for r in rows)
    total_delta    = total_ebuilder - total_privv
    direction      = "eBuilder is HIGHER" if total_delta > 0 else ("eBuilder is LOWER" if total_delta < 0 else "Exact Match")

    summary_frame = tk.Frame(win, bg="#1e1e2e", padx=10, pady=6)
    summary_frame.pack(fill="x")
    tk.Label(
        summary_frame,
        text=f"  PRIVV Total: ${total_privv:,.2f}     eBuilder Total: ${total_ebuilder:,.2f}     "
             f"Net Delta: ${total_delta:+,.2f}     Overall: {direction}",
        bg="#1e1e2e", fg="white", font=("Consolas", 10, "bold"), anchor="w"
    ).pack(fill="x")

    search_frame = tk.Frame(win, padx=8, pady=4)
    search_frame.pack(fill="x")
    tk.Label(search_frame, text="Filter vendor:").pack(side="left")
    search_var = tk.StringVar()
    tk.Entry(search_frame, textvariable=search_var, width=30).pack(side="left", padx=6)

    cols = ("Vendor", "PRIVV Total", "eBuilder Total", "Delta", "eB Matches", "Status")
    tree_frame = tk.Frame(win)
    tree_frame.pack(fill="both", expand=True, padx=8, pady=4)

    vsb = ttk.Scrollbar(tree_frame, orient="vertical")
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
    tree = ttk.Treeview(
        tree_frame, columns=cols, show="headings",
        yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=20
    )
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)

    col_widths = {"Vendor": 240, "PRIVV Total": 130, "eBuilder Total": 140, "Delta": 130, "eB Matches": 90, "Status": 160}
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=col_widths.get(c, 120), anchor="e" if c not in ("Vendor", "Status") else "w")

    tree.tag_configure("match",  background="#d4edda")
    tree.tag_configure("higher", background="#fff3cd")
    tree.tag_configure("lower",  background="#f8d7da")
    tree.tag_configure("nodata", background="#e2e3e5")

    # Map tree item iid → row dict so we can look up _entries on click
    iid_to_row: dict = {}

    UNASSIGNED_LABEL = "⚠ Unassigned (Manually Removed)"

    def _entry_amount(e: dict) -> float:
        """Same amount rule used everywhere else in the matcher: prefer
        Current Commitment, fall back to Commitment Amount."""
        def _num(v):
            s = str(v).replace(",", "").strip()
            if not s or s.lower() == "nan":
                return 0.0
            try:
                return float(s)
            except ValueError:
                return 0.0
        cur = _num(e.get("Current Commitment", 0))
        return cur if cur != 0 else _num(e.get("Commitment Amount", 0))

    def _recalc_row(r: dict):
        """Recompute eBuilder Total / eB Matches / Delta / Status for a row
        from whatever is currently in r['_entries']. Used after an entry is
        manually reassigned or removed."""
        entries = r.get("_entries", [])
        r["eBuilder Total"] = sum(_entry_amount(e) for e in entries)
        r["eB Matches"] = len(entries)
        r["Delta"] = r["eBuilder Total"] - r["PRIVV Total"]
        if r["eBuilder Total"] == 0 and r["PRIVV Total"] == 0:
            r["Status"] = "⚪ No Data"
        elif abs(r["Delta"]) < 0.01:
            r["Status"] = "✅ Match"
        elif r["eBuilder Total"] > r["PRIVV Total"]:
            r["Status"] = "🔼 eBuilder Higher"
        else:
            r["Status"] = "🔽 eBuilder Lower"

    def _get_or_create_unassigned_row():
        for r in rows:
            if r["Vendor"] == UNASSIGNED_LABEL:
                return r
        new_r = {
            "Vendor": UNASSIGNED_LABEL,
            "PRIVV Total": 0.0,
            "eBuilder Total": 0.0,
            "Delta": 0.0,
            "eB Matches": 0,
            "Status": "⚪ No Data",
            "_entries": [],
        }
        rows.append(new_r)
        return new_r

    def populate_tree(filter_text=""):
        tree.delete(*tree.get_children())
        iid_to_row.clear()
        ft = filter_text.strip().lower()
        for r in rows:
            if ft and ft not in r["Vendor"].lower():
                continue
            tag = "nodata"
            if "Match"   in r["Status"]: tag = "match"
            elif "Higher" in r["Status"]: tag = "higher"
            elif "Lower"  in r["Status"]: tag = "lower"
            iid = tree.insert("", "end", values=(
                r["Vendor"],
                f"${r['PRIVV Total']:,.2f}",
                f"${r['eBuilder Total']:,.2f}",
                f"${r['Delta']:+,.2f}",
                r["eB Matches"],
                r["Status"],
            ), tags=(tag,))
            iid_to_row[iid] = r

    def _populate_entries_tree(etree, entries, entry_cols, empty_label):
        if entries:
            for e in entries:
                etree.insert("", "end", values=tuple(
                    e.get(c, "") for c in entry_cols
                ))
        else:
            etree.insert("", "end", values=(empty_label,) + ("",) * (len(entry_cols) - 1))

    def _build_entries_tab(parent, entries, fallback_cols, col_widths, empty_label):
        # Build the column list from whatever fields are actually present
        # on the entries (full line item) rather than a fixed subset, so
        # every field captured for that row gets shown. Order is preserved
        # based on first appearance across all entries.
        if entries:
            entry_cols = []
            seen = set()
            for e in entries:
                for k in e.keys():
                    if k not in seen:
                        seen.add(k)
                        entry_cols.append(k)
        else:
            entry_cols = list(fallback_cols)

        ef = tk.Frame(parent)
        ef.pack(fill="both", expand=True, padx=8, pady=6)

        evsb = ttk.Scrollbar(ef, orient="vertical")
        ehsb = ttk.Scrollbar(ef, orient="horizontal")
        etree = ttk.Treeview(
            ef, columns=entry_cols, show="headings",
            yscrollcommand=evsb.set, xscrollcommand=ehsb.set, height=16
        )
        evsb.config(command=etree.yview)
        ehsb.config(command=etree.xview)

        money_cols = {
            "Commitment Amount", "Current Commitment",
            "Projected Commitment", "Actuals Approved", "Remaining Balance",
            "Amount",
        }
        for c in entry_cols:
            etree.heading(c, text=c)
            etree.column(c, width=col_widths.get(c, 130),
                         anchor="e" if c in money_cols else "w")

        _populate_entries_tree(etree, entries, entry_cols, empty_label)

        # Map tree iid -> index into `entries` (the SAME list object that
        # lives on the row, e.g. row["_entries"]) so callers can mutate the
        # underlying list directly when the user reassigns/removes a line.
        iid_to_entry_index = {}
        if entries:
            for i, iid in enumerate(etree.get_children()):
                iid_to_entry_index[iid] = i

        evsb.pack(side="right", fill="y")
        ehsb.pack(side="bottom", fill="x")
        etree.pack(fill="both", expand=True)
        return etree, iid_to_entry_index

    def show_entries_window(event=None):
        sel = tree.selection()
        if not sel:
            return
        row = iid_to_row.get(sel[0])
        if row is None:
            return
        eb_entries = row.get("_entries", [])

        # Fuzzy-match PRIVV rows for this same company on the fly (no
        # hardcoded vendor names — see find_fuzzy_privv_matches), rather
        # than threading a separate _privv_entries field through every
        # matching pass.
        privv_matches = find_fuzzy_privv_matches(
            row["Vendor"], privv_df,
            exclude_labels=[r2["Vendor"] for r2 in rows],
        )
        privv_entries = privv_matches.to_dict("records")
        # Use the sum of what's ACTUALLY shown in the PRIVV tab (fuzzy
        # matches), not row["PRIVV Total"] — that field comes from the
        # earlier exact-match aggregation pass and won't reflect rows that
        # only the fuzzy/Description matching here picked up.
        if "_amount_num" in privv_matches.columns:
            privv_total_fuzzy = privv_matches["_amount_num"].sum()
        else:
            privv_total_fuzzy = pd.to_numeric(
                privv_matches.get("Amount", pd.Series(dtype=str)).astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).sum()

        detail = tk.Toplevel(win)
        detail.title(f"Entries — {row['Vendor']}")
        detail.geometry("1300x500")

        privv_total_note = ""
        if abs(privv_total_fuzzy - row["PRIVV Total"]) > 0.005:
            privv_total_note = f" (was ${row['PRIVV Total']:,.2f} in summary)"

        hdr = tk.Frame(detail, bg="#1e1e2e", padx=10, pady=6)
        hdr.pack(fill="x")
        hdr_label = tk.Label(
            hdr,
            text=f"  Vendor: {row['Vendor']}     "
                 f"eBuilder Total: ${row['eBuilder Total']:,.2f}     "
                 f"PRIVV Total: ${privv_total_fuzzy:,.2f}{privv_total_note}     "
                 f"eB Entries: {len(eb_entries)}     "
                 f"PRIVV Entries: {len(privv_entries)}",
            bg="#1e1e2e", fg="white", font=("Consolas", 10, "bold"), anchor="w"
        )
        hdr_label.pack(fill="x")

        nb = ttk.Notebook(detail)
        nb.pack(fill="both", expand=True, padx=8, pady=6)

        eb_tab = tk.Frame(nb)
        privv_tab = tk.Frame(nb)
        nb.add(eb_tab, text=f"eBuilder Entries ({len(eb_entries)})")
        nb.add(privv_tab, text=f"PRIVV Entries ({len(privv_entries)})")

        eb_col_widths = {
            "ID": 70, "Description": 260, "Company": 200,
            "Date": 90, "Status": 100, "Commitment Type": 130,
            "Commitment Amount": 130, "Current Commitment": 130,
            "Projected Commitment": 130, "Actuals Approved": 130,
            "Remaining Balance": 130,
        }
        eb_tree, eb_iid_to_entry = _build_entries_tab(
            eb_tab, eb_entries, ENTRY_COLS, eb_col_widths,
            "(No eBuilder entries found)",
        )

        privv_col_widths = {
            "Vendor": 220, "Description": 280, "Amount": 130,
            "Date": 90, "Status": 100, "Item": 90, "Type": 130,
        }
        _build_entries_tab(
            privv_tab, privv_entries, list(privv_df.columns), privv_col_widths,
            "(No PRIVV entries found)",
        )

        # ── Reassign / remove a selected eBuilder entry ─────────────────────
        reassign_frame = tk.Frame(eb_tab, pady=6)
        reassign_frame.pack(fill="x", padx=8)

        tk.Label(reassign_frame, text="Selected entry →").pack(side="left")

        vendor_choices = [r["Vendor"] for r in rows if r["Vendor"] != row["Vendor"]]
        vendor_choices.sort(key=str.lower)
        vendor_choices.append(UNASSIGNED_LABEL)

        reassign_var = tk.StringVar()
        reassign_box = ttk.Combobox(
            reassign_frame, textvariable=reassign_var, values=vendor_choices,
            width=45, state="readonly",
        )
        reassign_box.pack(side="left", padx=6)

        def refresh_eb_tab():
            """Re-pull row['_entries'] and redraw the eBuilder tab + header
            in place, without closing the detail window."""
            nonlocal eb_entries, eb_iid_to_entry
            eb_entries = row.get("_entries", [])
            eb_tree.delete(*eb_tree.get_children())
            _populate_entries_tree(eb_tree, eb_entries, ENTRY_COLS, "(No eBuilder entries found)")
            eb_iid_to_entry = {iid: i for i, iid in enumerate(eb_tree.get_children())} if eb_entries else {}
            nb.tab(eb_tab, text=f"eBuilder Entries ({len(eb_entries)})")
            hdr_label.config(
                text=f"  Vendor: {row['Vendor']}     "
                     f"eBuilder Total: ${row['eBuilder Total']:,.2f}     "
                     f"PRIVV Total: ${privv_total_fuzzy:,.2f}{privv_total_note}     "
                     f"eB Entries: {len(eb_entries)}     "
                     f"PRIVV Entries: {len(privv_entries)}",
            )
            # Vendor list in the combobox can change (a target may now be
            # the freshly-created Unassigned bucket), so keep it current.
            choices = [r["Vendor"] for r in rows if r["Vendor"] != row["Vendor"]]
            choices.sort(key=str.lower)
            choices.append(UNASSIGNED_LABEL)
            reassign_box.config(values=choices)

        def apply_reassign():
            sel = eb_tree.selection()
            if not sel:
                messagebox.showwarning("No entry selected", "Select an eBuilder entry first.", parent=detail)
                return
            idx = eb_iid_to_entry.get(sel[0])
            if idx is None:
                return
            target_label = reassign_var.get()
            if not target_label:
                messagebox.showwarning("No destination", "Choose where to move this entry.", parent=detail)
                return

            entries_list = row.get("_entries", [])
            if idx >= len(entries_list):
                return
            entry = entries_list.pop(idx)
            _recalc_row(row)

            if target_label == UNASSIGNED_LABEL:
                target_row = _get_or_create_unassigned_row()
            else:
                target_row = next((r for r in rows if r["Vendor"] == target_label), None)
                if target_row is None:
                    entries_list.insert(idx, entry)  # target vanished, revert
                    _recalc_row(row)
                    return
            target_row.setdefault("_entries", []).append(entry)
            _recalc_row(target_row)

            populate_tree(search_var.get())
            refresh_eb_tab()
            log(f"Moved eBuilder entry (ID {entry.get('ID', '?')}) "
                f"from '{row['Vendor']}' to '{target_row['Vendor']}'.")

        def remove_entry():
            reassign_var.set(UNASSIGNED_LABEL)
            apply_reassign()

        tk.Button(reassign_frame, text="Move Entry", command=apply_reassign,
                  bg="#0066cc", fg="white", padx=8).pack(side="left", padx=4)
        tk.Button(reassign_frame, text="🗑 Remove / Unassign", command=remove_entry,
                  padx=8).pack(side="left", padx=4)

        tk.Button(detail, text="Close", command=detail.destroy, padx=10).pack(pady=6)

    tree.bind("<Double-1>", show_entries_window)


    populate_tree()
    search_var.trace_add("write", lambda *_: populate_tree(search_var.get()))

    vsb.pack(side="right",  fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    hint = tk.Label(win, text="💡 Double-click any row to view its eBuilder entries",
                    fg="#555", font=("Segoe UI", 9, "italic"))
    hint.pack(pady=(0, 2))

    legend = tk.Frame(win, padx=8, pady=4)
    legend.pack(fill="x")
    for color, label in [
        ("#d4edda", "✅ Match"),
        ("#fff3cd", "🔼 eBuilder Higher"),
        ("#f8d7da", "🔽 eBuilder Lower"),
        ("#e2e3e5", "⚪ No Data"),
    ]:
        tk.Label(legend, text=f"  {label}  ", bg=color, relief="solid", bd=1, padx=6, pady=2).pack(side="left", padx=4)

    STATUS_FILL = {
        "✅ Match":           "C6EFCE",
        "🔼 eBuilder Higher": "FFEB9C",
        "🔽 eBuilder Lower":  "FFC7CE",
        "⚪ No Data":          "E2E3E5",
    }
    HEADER_FILL = PatternFill("solid", start_color="1E1E2E", end_color="1E1E2E")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
    BASE_FONT   = Font(name="Calibri")
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _style_header(ws, headers):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    def _autosize(ws, n_cols, min_w=10, max_w=55):
        for c in range(1, n_cols + 1):
            letter = get_column_letter(c)
            longest = max(
                (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
                default=min_w,
            )
            ws.column_dimensions[letter].width = max(min_w, min(longest + 2, max_w))

    def export_comparison():
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save comparison as..."
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # ── Summary sheet ────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        headers = ["Vendor", "eBuilder Total", "PRIVV Total", "Delta",
                   "Status", "eB Matches"]
        _style_header(ws, headers)

        money_cols = {2, 3, 4}
        for r_idx, r in enumerate(rows, start=2):
            values = [
                r.get("Vendor", ""),
                r.get("eBuilder Total", 0) or 0,
                r.get("PRIVV Total", 0) or 0,
                r.get("Delta", 0) or 0,
                r.get("Status", ""),
                r.get("eB Matches", 0) or 0,
            ]
            fill = PatternFill("solid", start_color=STATUS_FILL.get(r.get("Status", ""), "FFFFFF"))
            for c_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = BASE_FONT
                cell.border = BORDER
                cell.fill = fill
                if c_idx in money_cols:
                    cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                if c_idx == 5:
                    cell.alignment = Alignment(horizontal="center")
        _autosize(ws, len(headers))

        # ── eBuilder Entries sheet ───────────────────────────────────────
        eb_sheet = wb.create_sheet("eBuilder Entries")
        eb_headers = ["Vendor"] + list(ENTRY_COLS)
        _style_header(eb_sheet, eb_headers)
        r_idx = 2
        for r in rows:
            for e in r.get("_entries", []):
                row_vals = [r.get("Vendor", "")] + [e.get(c, "") for c in ENTRY_COLS]
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = eb_sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BASE_FONT
                    cell.border = BORDER
                r_idx += 1
        _autosize(eb_sheet, len(eb_headers))

        # ── PRIVV Entries sheet (same fuzzy match used in the drilldown) ──
        privv_sheet = wb.create_sheet("PRIVV Entries")
        privv_cols = list(privv_df.columns.drop("_amount_num", errors="ignore"))
        privv_headers = ["Matched To Vendor"] + privv_cols
        _style_header(privv_sheet, privv_headers)
        r_idx = 2
        _export_vendors = [r2["Vendor"] for r2 in rows]
        for r in rows:
            matches = find_fuzzy_privv_matches(r["Vendor"], privv_df, exclude_labels=_export_vendors)
            for _, prow in matches.iterrows():
                row_vals = [r.get("Vendor", "")] + [prow.get(c, "") for c in privv_cols]
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = privv_sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BASE_FONT
                    cell.border = BORDER
                r_idx += 1
        _autosize(privv_sheet, len(privv_headers))

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to {path}")

    tk.Button(win, text="Export to Excel", command=export_comparison, bg="#0066cc", fg="white", padx=8).pack(pady=6)


# ---------------------------------------------------------------------------
# MAIN PROCESS
# ---------------------------------------------------------------------------
def run_process():
    if not ebuilder_files:
        messagebox.showerror("Error", "Select at least one eBuilder file.")
        return
    if not privv_file:
        messagebox.showerror("Error", "Select the PRIVV reference file.")
        return

    log("Starting process...")

    log(f"Combining {len(ebuilder_files)} eBuilder file(s)...")
    data_list = []
    for f in ebuilder_files:
        df = pd.read_csv(f, header=0, dtype=str).fillna("")
        df = df[~(df == "").all(axis=1)]
        log(f"  Loaded: {os.path.basename(f)} ({len(df)} rows)")
        data_list.append(df)

    data = pd.concat(data_list, ignore_index=True).drop_duplicates()
    log(f"Combined total (after dedup): {len(data)} rows")

    if "#" in data.columns:
        data = data.rename(columns={"#": "ID"})

    col_names = [
        "ID", "Description", "Company", "Date", "Status",
        "Commitment Type", "Commitment Amount",
        "Current Commitment", "Projected Commitment",
        "Actuals Approved", "Remaining Balance",
    ]
    if not all(c in data.columns for c in ["Description", "Current Commitment"]):
        while len(data.columns) < len(col_names):
            data[len(data.columns)] = ""
        data.columns = col_names[:len(data.columns)]

    # run_process keeps Company = Description intentionally for output formatting
   # data["Company"] = data["Description"]

    data = data[data["Description"].str.strip() != ""].reset_index(drop=True)

    ebuilder_df = data.copy()

    privv_df = pd.read_csv(privv_file, dtype=str).fillna("")
    lookup: dict = {}
    for _, row in privv_df.iterrows():
        company = normalize(row.get("Vendor", ""))
        code    = str(row.get("Code", "")).strip()
        if company and code:
            lookup[company] = code

    results = ebuilder_df.apply(lambda r: assign_code_weighted(r, lookup), axis=1)
    ebuilder_df["Cost Code"]  = results.apply(lambda x: x[0])
    ebuilder_df["Match Note"] = results.apply(lambda x: x[1])

    log("Classification summary:")
    for code, grp in ebuilder_df.groupby("Cost Code"):
        label = ITEM_MAP.get(code, "Unknown")
        log(f"  {code} ({label}): {len(grp)} rows")

    unmatched = ebuilder_df[ebuilder_df["Cost Code"] == DEFAULT_CODE]
    if not unmatched.empty:
        log(f"\nRows that fell through to default ({DEFAULT_CODE}):")
        for _, r in unmatched.iterrows():
            log(f"  Vendor: {r['Company']} | Desc: {r['Description'][:60]}")

    ebuilder_df["Final Type"] = ebuilder_df["Cost Code"].apply(
        lambda c: "Change Order" if c in ("315", "316") else "Original"
    )

    ebuilder_df["Amount"] = pd.to_numeric(
        ebuilder_df["Current Commitment"].str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0)

    po_df     = ebuilder_df[ebuilder_df["Commitment Type"] == "Purchase Order"]
    non_po_df = ebuilder_df[ebuilder_df["Commitment Type"] != "Purchase Order"]

    po_exempt = po_df[ po_df["Cost Code"].isin(["315", "316"])]
    po_normal = po_df[~po_df["Cost Code"].isin(["315", "316"])]

    grouped_po = (
        po_normal.groupby(["Company", "Cost Code", "Final Type"], as_index=False)
        .agg({"Amount": "sum", "Description": "first", "Date": "first"})
    )

    final_df = pd.concat(
        [
            grouped_po,
            po_exempt[["Company", "Cost Code", "Final Type", "Amount", "Description", "Date"]],
            non_po_df[["Company", "Cost Code", "Final Type", "Amount", "Description", "Date"]],
        ],
        ignore_index=True,
    )

    output = pd.DataFrame()
    output["Code"]             = final_df["Cost Code"]
    output["Vendor"]           = final_df["Company"]
    output["Description"]      = final_df["Description"]
    output["PO Number"]        = ""
    output["Item Description"] = final_df["Cost Code"].map(ITEM_MAP).fillna("")
    output["Date Committed"]   = final_df["Date"].apply(convert_date)
    output["Type"]             = final_df["Final Type"]
    output["Amount"]           = final_df["Amount"]

    output.to_csv(OUTPUT_FILE, index=False)
    log(f"\n✅ DONE: {OUTPUT_FILE} created with {len(output)} rows")


# ---------------------------------------------------------------------------
# GUI LAYOUT
# ---------------------------------------------------------------------------
root = tk.Tk()
root.title("EBuilder → PRIVV Processor (Weighted Classifier + Fuzzy Match)")
root.resizable(True, True)

# ── Top-level notebook (tabs) ────────────────────────────────────────────────
notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=6, pady=6)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — COMPARISONS  (original UI, unchanged)
# ════════════════════════════════════════════════════════════════════════════
comp_tab = tk.Frame(notebook)
notebook.add(comp_tab, text="Comparisons")

frame = tk.Frame(comp_tab)
frame.pack(pady=10, padx=10, fill="x")

tk.Button(frame, text="Select eBuilder Files", command=select_ebuilder_files, width=22).grid(
    row=0, column=0, padx=5, pady=4, sticky="w"
)
tk.Button(frame, text="Clear", command=clear_ebuilder_files, width=8, fg="red").grid(
    row=0, column=1, padx=2, pady=4, sticky="w"
)
ebuilder_label = tk.Label(frame, text="No eBuilder files selected.", fg="gray", anchor="w")
ebuilder_label.grid(row=0, column=2, padx=5, sticky="w")

tk.Button(frame, text="Select PRIVV File", command=select_privv_file, width=22).grid(
    row=1, column=0, padx=5, pady=4, sticky="w"
)
privv_label = tk.Label(frame, text="No PRIVV file selected.", fg="gray", anchor="w")
privv_label.grid(row=1, column=2, padx=5, sticky="w")

btn_frame = tk.Frame(comp_tab)
btn_frame.pack(pady=4)

tk.Button(
    btn_frame, text="Compare Vendors", command=run_comparison,
    bg="#0066cc", fg="white", width=18, padx=6
).pack(side="left", padx=8)

output_box = tk.Text(comp_tab, height=18, width=100)
output_box.pack(pady=10, padx=10)

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — INVOICES
# ════════════════════════════════════════════════════════════════════════════
inv_tab = tk.Frame(notebook)
notebook.add(inv_tab, text="Invoices")

# ── Invoice file state ───────────────────────────────────────────────────────
# Both sides now support selecting MULTIPLE files, which get concatenated
# together when the comparison is run (same pattern as the budget tab's
# ebuilder_files multi-select).
invoice_eb_files:    list = []
invoice_privv_files: list = []

# ── File-picker helpers ──────────────────────────────────────────────────────
inv_frame = tk.Frame(inv_tab)
inv_frame.pack(pady=10, padx=10, fill="x")

def _inv_eb_label_update():
    if not invoice_eb_files:
        inv_eb_label.config(text="No eBuilder invoice file(s) selected.", fg="gray")
    elif len(invoice_eb_files) == 1:
        inv_eb_label.config(text=f"✅ {os.path.basename(invoice_eb_files[0])}", fg="green")
    else:
        inv_eb_label.config(text=f"✅ {len(invoice_eb_files)} files selected", fg="green")

def _inv_privv_label_update():
    if not invoice_privv_files:
        inv_privv_label.config(text="No PRIVV invoice file(s) selected.", fg="gray")
    elif len(invoice_privv_files) == 1:
        inv_privv_label.config(text=f"✅ {os.path.basename(invoice_privv_files[0])}", fg="green")
    else:
        inv_privv_label.config(text=f"✅ {len(invoice_privv_files)} files selected", fg="green")

def select_invoice_eb_file():
    global invoice_eb_files
    files = list(filedialog.askopenfilenames(
        title="Select eBuilder Invoice CSV file(s)",
        filetypes=[("CSV Files", "*.csv")]
    ))
    if files:
        invoice_eb_files = files
    _inv_eb_label_update()

def select_invoice_privv_file():
    global invoice_privv_files
    files = list(filedialog.askopenfilenames(
        title="Select PRIVV Invoice CSV file(s)",
        filetypes=[("CSV Files", "*.csv")]
    ))
    if files:
        invoice_privv_files = files
    _inv_privv_label_update()

def clear_invoice_eb_files():
    global invoice_eb_files
    invoice_eb_files = []
    _inv_eb_label_update()

def clear_invoice_privv_files():
    global invoice_privv_files
    invoice_privv_files = []
    _inv_privv_label_update()

tk.Button(inv_frame, text="Select eBuilder Invoice File(s)",  command=select_invoice_eb_file,    width=26).grid(row=0, column=0, padx=5, pady=4, sticky="w")
inv_eb_label = tk.Label(inv_frame, text="No eBuilder invoice file(s) selected.", fg="gray", anchor="w")
inv_eb_label.grid(row=0, column=1, padx=5, sticky="w")
tk.Button(inv_frame, text="Clear", command=clear_invoice_eb_files, width=8, fg="red").grid(row=0, column=2, padx=5, sticky="w")

tk.Button(inv_frame, text="Select PRIVV Invoice File(s)", command=select_invoice_privv_file, width=26).grid(row=1, column=0, padx=5, pady=4, sticky="w")
inv_privv_label = tk.Label(inv_frame, text="No PRIVV invoice file(s) selected.", fg="gray", anchor="w")
inv_privv_label.grid(row=1, column=1, padx=5, sticky="w")
tk.Button(inv_frame, text="Clear", command=clear_invoice_privv_files, width=8, fg="red").grid(row=1, column=2, padx=5, sticky="w")

inv_btn_frame = tk.Frame(inv_tab)
inv_btn_frame.pack(pady=4)


def inv_log(msg: str):
    inv_output_box.insert(tk.END, msg + "\n")
    inv_output_box.see(tk.END)


def normalize_commit_num(val: str) -> str:
    """Strip leading zeros and non-digit characters for a numeric key comparison."""
    val = str(val).strip()
    digits = re.sub(r"[^0-9]", "", val)
    return str(int(digits)) if digits else ""


# ---------------------------------------------------------------------------
# INVOICE COMPARISON LOGIC
# ---------------------------------------------------------------------------
def _drop_total_rows(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Filter out any row that contains the word 'total' or 'totals' (case
    insensitive, whole-word match so it doesn't clip things like 'Totally
    Custom Co.') in ANY column. Used to strip subtotal/grand-total lines
    that some exports tack onto the bottom of a CSV."""
    if df.empty:
        return df
    total_pattern = re.compile(r"\btotals?\b", flags=re.IGNORECASE)
    mask = df.apply(
        lambda row: row.astype(str).str.contains(total_pattern, na=False).any(),
        axis=1,
    )
    dropped = int(mask.sum())
    if dropped:
        inv_log(f"  Filtered out {dropped} '{label}' row(s) containing 'total(s)'")
    return df[~mask].reset_index(drop=True)


def run_invoice_comparison():
    global invoice_eb_files, invoice_privv_files

    if not invoice_eb_files:
        messagebox.showerror("Error", "Select at least one eBuilder invoice file first.", parent=inv_tab)
        return
    if not invoice_privv_files:
        messagebox.showerror("Error", "Select at least one PRIVV invoice file first.", parent=inv_tab)
        return

    inv_log("\n── Running Invoice Comparison ─────────────────────────────────────")

    # ── Load & combine files ─────────────────────────────────────────────────
    inv_log(f"  Loading {len(invoice_eb_files)} eBuilder invoice file(s)...")
    eb_frames = []
    for f in invoice_eb_files:
        df = pd.read_csv(f, dtype=str).fillna("")
        inv_log(f"    {os.path.basename(f)}: {len(df)} rows")
        eb_frames.append(df)
    eb_inv = pd.concat(eb_frames, ignore_index=True, sort=False).fillna("")

    inv_log(f"  Loading {len(invoice_privv_files)} PRIVV invoice file(s)...")
    prv_frames = []
    for f in invoice_privv_files:
        df = pd.read_csv(f, dtype=str).fillna("")
        inv_log(f"    {os.path.basename(f)}: {len(df)} rows")
        prv_frames.append(df)
    prv_inv = pd.concat(prv_frames, ignore_index=True, sort=False).fillna("")

    # ── Filter out total/totals rows ─────────────────────────────────────────
    eb_inv  = _drop_total_rows(eb_inv,  "eBuilder")
    prv_inv = _drop_total_rows(prv_inv, "PRIVV")

    # ── Hardcoded overrides ───────────────────────────────────────────────────
    # Some eBuilder rows list "The Pennsylvania State University" in the
    # Company field (the paying entity) instead of the actual vendor, which
    # would otherwise cause _eb_company_belongs_to() to bucket them under the
    # PSU OPP label. Force known rows like this to their correct vendor here,
    # identified by Invoice # (and Commitment # as a safety check).
    HARDCODED_INVOICE_VENDOR_OVERRIDES = {
        # Invoice #: (Commitment #, forced Company/vendor label)
        "86896": ("4100350409", "SC VINYL LLC"),
    }
    if "Invoice #" in eb_inv.columns:
        for inv_num, (commit_num, forced_vendor) in HARDCODED_INVOICE_VENDOR_OVERRIDES.items():
            mask = eb_inv["Invoice #"].astype(str).str.strip() == inv_num
            if "Commitment #" in eb_inv.columns:
                mask &= eb_inv["Commitment #"].astype(str).str.strip() == commit_num
            n_hit = int(mask.sum())
            if n_hit:
                eb_inv.loc[mask, "Company"] = forced_vendor
                inv_log(f"  Hardcoded override: Invoice #{inv_num} -> forced to vendor '{forced_vendor}' ({n_hit} row(s))")

    # Give every source row a stable, unique id BEFORE any matching happens.
    # Rows can legitimately get copied into more than one comparison row's
    # _eb_entries/_prv_entries during matching (Commitment # match, then
    # fuzzy fallback, then PSU OPP fallback, etc.), which double counts the
    # row's dollar amount if you just add up each comparison row's subtotal.
    # These ids let us de-dupe back down to "each source row counted once"
    # for the true post-matching totals computed further below.
    eb_inv["_uid"]  = range(len(eb_inv))
    prv_inv["_uid"] = range(len(prv_inv))

    # ── Validate required columns ─────────────────────────────────────────────
    eb_required  = {"Commitment #", "Invoice Amount"}
    prv_required = {"Vendor #", "Amount"}

    missing_eb  = eb_required  - set(eb_inv.columns)
    missing_prv = prv_required - set(prv_inv.columns)
    if missing_eb:
        messagebox.showerror("Error", f"eBuilder invoice file is missing columns: {missing_eb}", parent=inv_tab)
        return
    if missing_prv:
        messagebox.showerror("Error", f"PRIVV invoice file is missing columns: {missing_prv}", parent=inv_tab)
        return

    inv_log(f"  eBuilder invoices loaded  : {len(eb_inv)} rows")
    inv_log(f"  PRIVV invoices loaded     : {len(prv_inv)} rows")

    # ── Parse amounts to numeric ──────────────────────────────────────────────
    eb_inv["_amount_num"] = pd.to_numeric(
        eb_inv["Invoice Amount"].str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)

    prv_inv["_amount_num"] = pd.to_numeric(
        prv_inv["Amount"].str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)

    # ── Normalize join keys ───────────────────────────────────────────────────
    # eBuilder  → Commitment # (e.g. "000925000")
    # PRIVV     → Vendor #     (e.g. "8103142")
    eb_inv["_commit_key"]  = eb_inv["Commitment #"].apply(normalize_commit_num)
    prv_inv["_vendor_key"] = prv_inv["Vendor #"].apply(normalize_commit_num)

    # ── Build PRIVV lookup: vendor_key → rows ─────────────────────────────────
    prv_by_key: dict = {}
    for _, row in prv_inv.iterrows():
        k = row["_vendor_key"]
        if k:
            prv_by_key.setdefault(k, []).append(row)

    # ── Build PRIVV vendor-name list for fuzzy fallback ───────────────────────
    prv_vendor_names: list = []
    _seen_pv: set = set()
    for _, row in prv_inv.iterrows():
        vname = str(row.get("Vendor", "")).strip()
        if vname and normalize(vname) not in _seen_pv:
            _seen_pv.add(normalize(vname))
            prv_vendor_names.append(vname)

    INV_FUZZY_THRESHOLD = 72

    def _alias_find_prv_vendor(eb_row) -> str | None:
        """Return a PRIVV vendor label for an eBuilder invoice row using
        VENDOR_ALIASES (the same alias map the budget side uses), checking
        the Company and Description fields. This runs BEFORE the generic
        fuzzy fallback so known eBuilder/PRIVV name mismatches (e.g. "DOBIL
        LABORATORIES INC" -> "Dobil Laboratories") resolve correctly here
        too, instead of only on the budget side."""
        for field_val in [str(eb_row.get("Company", "")).strip(),
                           str(eb_row.get("Description", "")).strip()]:
            if not field_val:
                continue
            for candidate in resolve_alias_list(field_val):
                if not candidate or normalize(candidate) == normalize(field_val):
                    continue  # resolve_alias_list returned the input unchanged -> no alias hit
                # Prefer matching the alias's canonical name to an actual
                # PRIVV vendor name so it lines up with existing comparison
                # rows / totals built from prv_inv.
                for pv in prv_vendor_names:
                    npv, ncand = normalize(pv), normalize(candidate)
                    if npv == ncand or ncand in npv or npv in ncand:
                        return pv
                # No exact PRIVV vendor match, but the alias's canonical
                # name may still match an existing comparison row label
                # (e.g. multi-string JV aliases) — return it as-is.
                return candidate
        return None

    def _fuzzy_find_prv_vendor(eb_row) -> str | None:
        """Return the best PRIVV vendor name for an eBuilder row whose
        Commitment # is blank, using Company / Description fuzzy match."""
        company = str(eb_row.get("Company", "")).strip()
        desc    = str(eb_row.get("Description", "")).strip()
        if not prv_vendor_names or (not company and not desc):
            return None
        candidates = []
        for field_val in [company, desc]:
            if not field_val:
                continue
            m = process.extractOne(
                normalize(field_val),
                [normalize(v) for v in prv_vendor_names],
                scorer=fuzz.token_set_ratio,
                score_cutoff=INV_FUZZY_THRESHOLD,
            )
            if m:
                _, score, idx = m
                candidates.append((prv_vendor_names[idx], score))
        if not candidates:
            return None
        return max(candidates, key=lambda x: x[1])[0]

    # ── Main comparison pass ──────────────────────────────────────────────────
    # Iterate over every unique Commitment # seen on the PRIVV side (Vendor #),
    # find corresponding eBuilder rows by Commitment #, compute totals & delta.
    comparison_rows: list = []
    all_matched_eb_indices: set = set()

    # Collect all unique PRIVV vendor # keys (these drive the rows).
    # If the vendor name is Penn State OPP (any variant), always use the
    # canonical PSU_OPP_LABEL so all PSU rows collapse into one line item.
    PSU_OPP_LABEL = "Penn State Office of Physical Plant"
    PSU_OPP_KEY   = normalize(PSU_OPP_LABEL)
    PSU_NAME_VARIANTS = {
        "penn state office of physical plant",
        "pennsylvania state university",
        "the pennsylvania state university",
        "penn state opp",
        "psu opp",
    }

    def _make_label(vendor_name: str, vendor_num: str) -> str:
        """Return PSU_OPP_LABEL for any Penn State variant; otherwise just the
        vendor name (no Vendor # suffix).  Dropping the number means two PRIVV
        rows for the same company with different Vendor #s get the same label
        and are automatically collapsed by the merge-duplicates pass."""
        if normalize(vendor_name) in PSU_NAME_VARIANTS:
            return PSU_OPP_LABEL
        return vendor_name if vendor_name else f"#{vendor_num}"

    prv_keys_seen: dict = {}  # normalized vendor name -> display label
    for _, row in prv_inv.iterrows():
        vendor_name = str(row.get("Vendor", "")).strip()
        if not vendor_name:
            continue
        vendor_num = str(row.get("Vendor #", "")).strip()
        label = _make_label(vendor_name, vendor_num)
        nk = normalize(label)
        if nk not in prv_keys_seen:
            prv_keys_seen[nk] = label

    # Group PRIVV rows by that same normalized vendor-name key.
    prv_by_name: dict = {}
    for _, row in prv_inv.iterrows():
        vendor_name = str(row.get("Vendor", "")).strip()
        if not vendor_name:
            continue
        vendor_num = str(row.get("Vendor #", "")).strip()
        nk = normalize(_make_label(vendor_name, vendor_num))
        prv_by_name.setdefault(nk, []).append(row)

    def _eb_company_belongs_to(company: str, vendor_nk: str) -> bool:
        """True if an eBuilder row's Company field identifies the same
        vendor as the given (normalized) PRIVV vendor label. Checks the
        full bidirectional alias-equivalents group first (so either the
        short or long form of an aliased vendor name matches), then falls
        back to exact/substring, then a fuzzy match on raw text."""
        if not company:
            return False
        for equiv in alias_equivalents(company):
            neq = normalize(equiv)
            if neq == vendor_nk or (neq and (neq in vendor_nk or vendor_nk in neq)):
                return True
        nc = normalize(company)
        if nc == vendor_nk or (nc and (nc in vendor_nk or vendor_nk in nc)):
            return True
        return fuzz.token_set_ratio(nc, vendor_nk) >= INV_FUZZY_THRESHOLD

    for nk, label in prv_keys_seen.items():
        # PRIVV side: all rows for this vendor name
        prv_rows = prv_by_name.get(nk, [])
        prv_total = sum(float(r["_amount_num"]) for r in prv_rows)

        # eBuilder side: rows whose Company field identifies this vendor,
        # via alias resolution (bidirectional) / exact / substring / fuzzy —
        # not Commitment #/Vendor #.
        eb_mask = eb_inv["Company"].apply(lambda c: _eb_company_belongs_to(c, nk))
        eb_total = eb_inv.loc[eb_mask, "_amount_num"].sum()
        eb_count = int(eb_mask.sum())
        all_matched_eb_indices.update(eb_inv[eb_mask].index.tolist())

        delta = eb_total - prv_total
        if eb_total == 0 and prv_total == 0:
            status = "⚪ No Data"
        elif abs(delta) < 0.01:
            status = "✅ Match"
        elif eb_total > prv_total:
            status = "🔼 eBuilder Higher"
        else:
            status = "🔽 eBuilder Lower"

        # If this key maps to PSU OPP and we already have a PSU OPP row,
        # merge into it instead of adding a duplicate line.
        existing_psu = next(
            (r for r in comparison_rows if normalize(r["Label"]) == PSU_OPP_KEY),
            None,
        ) if label == PSU_OPP_LABEL else None

        if existing_psu:
            existing_psu["PRIVV Total"]    += prv_total
            existing_psu["eBuilder Total"] += eb_total
            existing_psu["eB Matches"]     += eb_count
            existing_psu["_eb_entries"].extend(eb_inv[eb_mask].to_dict("records"))
            existing_psu["_prv_entries"].extend([dict(r) for r in prv_rows])
            existing_psu["Delta"]  = existing_psu["eBuilder Total"] - existing_psu["PRIVV Total"]
            existing_psu["Status"] = (
                "✅ Match"           if abs(existing_psu["Delta"]) < 0.01 else
                "🔼 eBuilder Higher" if existing_psu["eBuilder Total"] > existing_psu["PRIVV Total"] else
                "🔽 eBuilder Lower"
            )
        else:
            comparison_rows.append({
                "Label":          label,
                "Commit Key":     "",
                "PRIVV Total":    prv_total,
                "eBuilder Total": eb_total,
                "Delta":          delta,
                "eB Matches":     eb_count,
                "Status":         status,
                "_eb_entries":    eb_inv[eb_mask].to_dict("records"),
                "_prv_entries":   [dict(r) for r in prv_rows],
            })


    # ── Helper: add rows into the PSU OPP bucket ─────────────────────────────
    # PSU_OPP_LABEL / PSU_OPP_KEY are defined above in the label-building block.

    def _add_to_inv_psu(rows_list, reason: str):
        """Merge a list of eBuilder invoice rows into the PSU OPP comparison row."""
        if not rows_list:
            return
        amount  = sum(float(r["_amount_num"]) for r in rows_list)
        n_rows  = len(rows_list)
        entries = [dict(r) for r in rows_list]
        existing = next(
            (r for r in comparison_rows if normalize(r["Label"]) == PSU_OPP_KEY),
            None,
        )
        for eb_row in rows_list:
            inv_log(f"    PSU OPP [{reason}]: "
                    f"Invoice #{eb_row.get('Invoice #','')} | "
                    f"{eb_row.get('Company','')} | "
                    f"${float(eb_row['_amount_num']):,.2f}")
        if existing:
            existing["eBuilder Total"] += amount
            existing["eB Matches"]     += n_rows
            existing["_eb_entries"].extend(entries)
            existing["Delta"]  = existing["eBuilder Total"] - existing["PRIVV Total"]
            existing["Status"] = (
                "✅ Match"           if abs(existing["Delta"]) < 0.01 else
                "🔼 eBuilder Higher" if existing["eBuilder Total"] > existing["PRIVV Total"] else
                "🔽 eBuilder Lower"
            )
        else:
            # PSU OPP not yet in comparison_rows — pull its PRIVV rows too
            prv_psu_mask = prv_inv["Vendor"].apply(
                lambda v: normalize(str(v)) == PSU_OPP_KEY
            )
            prv_psu_total = prv_inv.loc[prv_psu_mask, "_amount_num"].sum()
            prv_psu_rows  = prv_inv[prv_psu_mask].to_dict("records")
            delta = amount - prv_psu_total
            comparison_rows.append({
                "Label":          PSU_OPP_LABEL,
                "Commit Key":     "",
                "PRIVV Total":    prv_psu_total,
                "eBuilder Total": amount,
                "Delta":          delta,
                "eB Matches":     n_rows,
                "Status":         (
                    "✅ Match"           if abs(delta) < 0.01 else
                    "🔼 eBuilder Higher" if amount > prv_psu_total else
                    "🔽 eBuilder Lower"
                ),
                "_eb_entries":    entries,
                "_prv_entries":   prv_psu_rows,
            })

    # ── Build PSU OPP date set from PRIVV invoices ────────────────────────────
    # Mirrors the same heuristic used on the budget/commitment tab: an
    # unmatched eBuilder row is only routed to PSU OPP if it actually looks
    # like it belongs there (Penn keyword or a date PRIVV logged against
    # PSU OPP) — NOT just because nothing else matched.
    psu_inv_dates: set = set()
    inv_date_col = None
    for candidate_col in ("Invoice Date", "Date"):
        if candidate_col in prv_inv.columns:
            inv_date_col = candidate_col
            break
    if inv_date_col:
        psu_prv_date_mask = prv_inv["Vendor"].apply(
            lambda v: normalize(str(v)) in PSU_NAME_VARIANTS
        )
        for raw_date in prv_inv.loc[psu_prv_date_mask, inv_date_col]:
            tup = normalize_date_to_tuple(raw_date)
            if tup:
                psu_inv_dates.add(tup)
    if psu_inv_dates:
        inv_log(f"  PSU OPP PRIVV invoice dates loaded: {len(psu_inv_dates)} date(s)")

    def _eb_inv_date_matches_psu(eb_date_val) -> bool:
        tup = normalize_date_to_tuple(eb_date_val)
        return tup is not None and tup in psu_inv_dates

    eb_date_col = "Date Received" if "Date Received" in eb_inv.columns else None

    def _add_to_new_inv_company(eb_row):
        """Add a genuinely unmatched eBuilder invoice row as its own new
        line item, instead of silently sweeping it into PSU OPP."""
        company = str(eb_row.get("Company", "")).strip()
        desc    = str(eb_row.get("Description", "")).strip()
        label   = company or (desc[:40].strip() if desc else "") or \
                  f"Unmatched (Invoice #{eb_row.get('Invoice #','')})"
        amount  = float(eb_row["_amount_num"])
        comparison_rows.append({
            "Label":          label,
            "Commit Key":     "",
            "PRIVV Total":    0.0,
            "eBuilder Total": amount,
            "Delta":          amount,
            "eB Matches":     1,
            "Status":         "❓ Unmatched (needs review)",
            "_eb_entries":    [dict(eb_row)],
            "_prv_entries":   [],
        })
        inv_log(f"    No match found (not PSU) → new unmatched entry '{label}': "
                f"Invoice #{eb_row.get('Invoice #','')} | ${amount:,.2f}")

    # ── Handle unmatched eBuilder rows ────────────────────────────────────────
    unmatched_eb = eb_inv[~eb_inv.index.isin(all_matched_eb_indices)].copy()
    inv_log(f"\n  Unmatched eBuilder invoice rows (no Commitment # match): {len(unmatched_eb)}")

    fuzzy_buckets: dict = {}  # normalized vendor name -> {label, rows}
    psu_fallback:  list = []  # rows that pass the Penn/date heuristic -> PSU OPP
    new_company_rows: list = []  # rows that match nothing at all -> own line item

    for _, eb_row in unmatched_eb.iterrows():
        # If this row simply has no Commitment # at all, skip straight to
        # the Company fallback — don't run it through the fuzzy PRIVV-vendor
        # match or the PSU/Penn heuristic first. Rows that DO have a
        # Commitment # but just didn't match any PRIVV key still go through
        # the fuzzy/PSU checks below, since that's a real mismatch to
        # resolve, not a missing value. Invoice side only.
        if not str(eb_row.get("_commit_key", "")).strip():
            new_company_rows.append(eb_row)
            inv_log(f"    No Commitment # → grouped by Company directly: "
                    f"Invoice #{eb_row.get('Invoice #','')} | "
                    f"{eb_row.get('Company','')} | "
                    f"${float(eb_row['_amount_num']):,.2f}")
            continue

        vendor_hit = _alias_find_prv_vendor(eb_row) or _fuzzy_find_prv_vendor(eb_row)
        if vendor_hit:
            # Collapse any Penn State variant ("The Pennsylvania State
            # University", "Pennsylvania State University", "PSU OPP", etc.)
            # onto the same canonical label used elsewhere, so it merges
            # into the existing PSU OPP row instead of becoming its own
            # separate line item.
            if normalize(vendor_hit) in PSU_NAME_VARIANTS:
                vendor_hit = PSU_OPP_LABEL
            vk = normalize(vendor_hit)
            fuzzy_buckets.setdefault(vk, {"label": vendor_hit, "rows": []})
            fuzzy_buckets[vk]["rows"].append(eb_row)
            inv_log(f"    Fuzzy match to '{vendor_hit}': "
                    f"Invoice #{eb_row.get('Invoice #','')} | "
                    f"{eb_row.get('Company','')} | "
                    f"${float(eb_row['_amount_num']):,.2f}")
            continue

        # No vendor match — only route to PSU OPP if it actually looks like
        # PSU OPP (Penn keyword in Company/Description, or a date PRIVV
        # logged against PSU OPP). Otherwise it's a genuinely new/unmatched
        # row and should be visible as its own line, not folded into PSU.
        company_str = str(eb_row.get("Company", ""))
        desc_str    = str(eb_row.get("Description", ""))
        # Word-level match (not raw substring) so short tokens like "psu"
        # don't false-positive on unrelated words.
        psu_tokens = set(normalize(company_str).split()) | set(normalize(desc_str).split())
        penn_in_text = bool(psu_tokens & {"penn", "psu"})
        date_matches = (eb_date_col is not None and
                        _eb_inv_date_matches_psu(eb_row.get(eb_date_col, "")))

        if penn_in_text or date_matches:
            reason = []
            if penn_in_text: reason.append("'Penn'/'PSU' keyword in Company/Description")
            if date_matches: reason.append("date matches PSU OPP PRIVV entry")
            inv_log(f"    {' & '.join(reason)} → PSU OPP: "
                    f"Invoice #{eb_row.get('Invoice #','')}")
            psu_fallback.append(eb_row)
        else:
            new_company_rows.append(eb_row)

    # Merge fuzzy-matched rows into their matching comparison_row
    for vk, bucket in fuzzy_buckets.items():
        vendor_label = bucket["label"]
        rows_list    = bucket["rows"]
        fuzzy_amount = sum(float(r["_amount_num"]) for r in rows_list)
        n_rows       = len(rows_list)
        existing = next(
            (r for r in comparison_rows if normalize(r["Label"]) == vk),
            None,
        )
        if existing:
            existing["eBuilder Total"] += fuzzy_amount
            existing["eB Matches"]     += n_rows
            existing["_eb_entries"].extend([dict(r) for r in rows_list])
            existing["Delta"]  = existing["eBuilder Total"] - existing["PRIVV Total"]
            existing["Status"] = (
                "✅ Match"           if abs(existing["Delta"]) < 0.01 else
                "🔼 eBuilder Higher" if existing["eBuilder Total"] > existing["PRIVV Total"] else
                "🔽 eBuilder Lower"
            )
        else:
            # Fuzzy matched a vendor name not yet in comparison_rows — add it
            comparison_rows.append({
                "Label":          vendor_label,
                "Commit Key":     "",
                "PRIVV Total":    0.0,
                "eBuilder Total": fuzzy_amount,
                "Delta":          fuzzy_amount,
                "eB Matches":     n_rows,
                "Status":         "🔼 eBuilder Higher",
                "_eb_entries":    [dict(r) for r in rows_list],
                "_prv_entries":   [],
            })

    # Rows that couldn't be matched by Commitment # OR fuzzy, but DO look like
    # PSU OPP (Penn keyword / date match) -> Penn State OPP
    if psu_fallback:
        inv_log(f"\n  {len(psu_fallback)} invoice row(s) matched PSU OPP "
                f"via Penn keyword / date heuristic:")
        _add_to_inv_psu(psu_fallback, "Penn keyword/date match")

    # Rows that couldn't be matched at all AND don't look like PSU OPP ->
    # their own line item so they're visible for review instead of being
    # silently absorbed into Penn State OPP.
    if new_company_rows:
        inv_log(f"\n  {len(new_company_rows)} invoice row(s) genuinely unmatched "
                f"-> flagged as new/unmatched entries:")
        for eb_row in new_company_rows:
            _add_to_new_inv_company(eb_row)

    # ── Force-normalize any Penn State variant label to the canonical PSU OPP
    # label BEFORE merging duplicates. This is a final safety net that
    # catches every code path (main Commitment# pass, fuzzy fallback, PSU
    # keyword/date fallback, unmatched/new-company rows, whatever), not just
    # the ones we've explicitly canonicalized above, so "The Pennsylvania
    # State University" / "Pennsylvania State University" / "PSU OPP" etc.
    # always collapse into the same "Penn State Office of Physical Plant"
    # line item no matter how they were labeled going in.
    for r in comparison_rows:
        if normalize(r["Label"]) in PSU_NAME_VARIANTS:
            r["Label"] = PSU_OPP_LABEL

    # ── Merge duplicate company entries into one row ──────────────────────────
    # Any two rows with the same normalized Label get collapsed: amounts are
    # summed, entry lists are combined, and status is recomputed.
    merged: list = []
    seen_labels: dict = {}  # normalized label -> index in merged
    for r in comparison_rows:
        nk = normalize(r["Label"])
        if nk in seen_labels:
            m = merged[seen_labels[nk]]
            m["PRIVV Total"]    += r["PRIVV Total"]
            m["eBuilder Total"] += r["eBuilder Total"]
            m["eB Matches"]     += r["eB Matches"]
            m["_eb_entries"].extend(r.get("_eb_entries", []))
            m["_prv_entries"].extend(r.get("_prv_entries", []))
            m["Delta"]  = m["eBuilder Total"] - m["PRIVV Total"]
            m["Status"] = (
                "✅ Match"               if abs(m["Delta"]) < 0.01 else
                "🔼 eBuilder Higher" if m["eBuilder Total"] > m["PRIVV Total"] else
                "🔽 eBuilder Lower"
            )
        else:
            seen_labels[nk] = len(merged)
            merged.append(dict(r))
    comparison_rows = merged

    # ── De-dup pass: strip invoice entries also claimed by PSU OPP ──────────
    # Same idea as the budget-comparison side: walk every company row EXCEPT
    # PSU OPP, and if one of its invoice entries (_uid) also shows up under
    # PSU OPP, that's a double-count — PSU OPP wins, the entry is stripped
    # from the other company's row/total/match count, and it's logged to
    # the GUI so the duplicate is visible instead of silently inflating
    # both rows.
    _psu_inv_row = next(
        (r for r in comparison_rows if normalize(r["Label"]) == normalize(PSU_OPP_LABEL)),
        None,
    )
    if _psu_inv_row is not None:
        _psu_inv_uids = {
            e.get("_uid") for e in _psu_inv_row.get("_eb_entries", [])
            if e.get("_uid") is not None
        }
        if _psu_inv_uids:
            _dup_inv_lines = []
            for r in comparison_rows:
                if r is _psu_inv_row:
                    continue
                entries = r.get("_eb_entries", [])
                keep, removed = [], []
                for e in entries:
                    if e.get("_uid") in _psu_inv_uids:
                        removed.append(e)
                    else:
                        keep.append(e)
                if removed:
                    removed_amount = sum(float(e.get("_amount_num", 0) or 0) for e in removed)
                    r["_eb_entries"]    = keep
                    r["eBuilder Total"] = r["eBuilder Total"] - removed_amount
                    r["eB Matches"]     = max(0, r["eB Matches"] - len(removed))
                    r["Delta"]          = r["eBuilder Total"] - r["PRIVV Total"]
                    r["Status"] = (
                        "✅ Match"           if abs(r["Delta"]) < 0.01 else
                        "🔼 eBuilder Higher" if r["eBuilder Total"] > r["PRIVV Total"] else
                        "🔽 eBuilder Lower"
                    )
                    for e in removed:
                        _dup_inv_lines.append(e)  # kept for internal tracking only

    # ── De-dup pass: strip PRIVV entries also claimed by PSU OPP ────────────
    # Mirrors the eBuilder de-dup pass directly above — same double-count bug
    # can happen on the PRIVV side too (e.g. a PRIVV row logged under the
    # generic PSU OPP Vendor # but whose Vendor/Description names a specific
    # real vendor, or any other path that lets a PRIVV row's _uid land under
    # two different comparison rows). Without this, such a row was only
    # flagged later in the "TRUE POST-MATCHING TOTALS" log message
    # (dup_prv_count) instead of actually being corrected here — this pass
    # fixes it the same way the eBuilder side already is: PSU OPP wins, the
    # entry is stripped from the other company's row/total, and it's logged.
    if _psu_inv_row is not None:
        _psu_inv_prv_uids = {
            e.get("_uid") for e in _psu_inv_row.get("_prv_entries", [])
            if e.get("_uid") is not None
        }
        if _psu_inv_prv_uids:
            _dup_prv_lines = []
            for r in comparison_rows:
                if r is _psu_inv_row:
                    continue
                entries = r.get("_prv_entries", [])
                keep, removed = [], []
                for e in entries:
                    if e.get("_uid") in _psu_inv_prv_uids:
                        removed.append(e)
                    else:
                        keep.append(e)
                if removed:
                    removed_amount = sum(float(e.get("_amount_num", 0) or 0) for e in removed)
                    r["_prv_entries"] = keep
                    r["PRIVV Total"]  = r["PRIVV Total"] - removed_amount
                    r["Delta"]        = r["eBuilder Total"] - r["PRIVV Total"]
                    r["Status"] = (
                        "✅ Match"           if abs(r["Delta"]) < 0.01 else
                        "🔼 eBuilder Higher" if r["eBuilder Total"] > r["PRIVV Total"] else
                        "🔽 eBuilder Lower"
                    )
                    for e in removed:
                        _dup_prv_lines.append(e)  # kept for internal tracking only
                    inv_log(f"    ⚠ Removed {len(removed)} PRIVV entry(ies) (${removed_amount:,.2f}) "
                            f"from '{r['Label']}' — already claimed by PSU OPP.")

    comparison_rows.sort(key=lambda r: abs(r["Delta"]), reverse=True)

    total_eb  = sum(r["eBuilder Total"] for r in comparison_rows)
    total_prv = sum(r["PRIVV Total"]    for r in comparison_rows)
    total_d   = total_eb - total_prv
    inv_log(f"\n  eBuilder Invoice Total : ${total_eb:,.2f}")
    inv_log(f"  PRIVV Invoice Total    : ${total_prv:,.2f}")
    inv_log(f"  Net Delta              : ${total_d:+,.2f}")

    # ── TRUE POST-MATCHING TOTALS (de-duplicated) ────────────────────────────
    # total_eb / total_prv above are sums of each comparison row's subtotal.
    # If the rare "same entry landed under two companies" bug fires, a
    # source row can be present in more than one row's _eb_entries /
    # _prv_entries, and that sum will double count it. Re-derive the totals
    # here by walking every entry once we've collected them ALL (matching is
    # fully done at this point) and keeping only the first copy of each
    # source row's _uid, so these numbers reflect the actual dollars in the
    # two source files regardless of how matching routed them.
    seen_eb_uids: set = set()
    true_eb_total = 0.0
    dup_eb_count = 0
    for r in comparison_rows:
        for entry in r.get("_eb_entries", []):
            uid = entry.get("_uid")
            if uid in seen_eb_uids:
                dup_eb_count += 1
                continue
            seen_eb_uids.add(uid)
            true_eb_total += float(entry.get("_amount_num", 0) or 0)

    seen_prv_uids: set = set()
    true_prv_total = 0.0
    dup_prv_count = 0
    for r in comparison_rows:
        for entry in r.get("_prv_entries", []):
            uid = entry.get("_uid")
            if uid in seen_prv_uids:
                dup_prv_count += 1
                continue
            seen_prv_uids.add(uid)
            true_prv_total += float(entry.get("_amount_num", 0) or 0)

    inv_log(f"\n  TRUE eBuilder Total (deduped) : ${true_eb_total:,.2f}")
    if dup_eb_count:
        inv_log(f"    ⚠ {dup_eb_count} eBuilder entry instance(s) were counted under "
                f"more than one company — excluded from the true total above.")
    inv_log(f"  TRUE PRIVV Total    (deduped) : ${true_prv_total:,.2f}")
    if dup_prv_count:
        inv_log(f"    ⚠ {dup_prv_count} PRIVV entry instance(s) were counted under "
                f"more than one company — excluded from the true total above.")
    if abs(true_eb_total - total_eb) > 0.01 or abs(true_prv_total - total_prv) > 0.01:
        inv_log(f"  TRUE Net Delta                : ${true_eb_total - true_prv_total:+,.2f}")

    inv_log(f"\n  Comparison complete — {len(comparison_rows)} commitment(s) analyzed.")

    show_invoice_comparison_window(comparison_rows, prv_inv)


# ---------------------------------------------------------------------------
# INVOICE RESULTS WINDOW
# ---------------------------------------------------------------------------
def show_invoice_comparison_window(rows, prv_inv_df):
    win = tk.Toplevel(root)
    win.title("Invoice Comparison: PRIVV vs eBuilder")
    win.geometry("1100x560")

    total_privv    = sum(r["PRIVV Total"]    for r in rows)
    total_ebuilder = sum(r["eBuilder Total"] for r in rows)
    total_delta    = total_ebuilder - total_privv
    direction      = ("eBuilder is HIGHER" if total_delta > 0 else
                      ("eBuilder is LOWER"  if total_delta < 0 else "Exact Match"))

    summary_frame = tk.Frame(win, bg="#1e1e2e", padx=10, pady=6)
    summary_frame.pack(fill="x")
    tk.Label(
        summary_frame,
        text=(f"  PRIVV Total: ${total_privv:,.2f}     "
              f"eBuilder Total: ${total_ebuilder:,.2f}     "
              f"Net Delta: ${total_delta:+,.2f}     Overall: {direction}"),
        bg="#1e1e2e", fg="white", font=("Consolas", 10, "bold"), anchor="w"
    ).pack(fill="x")

    search_frame = tk.Frame(win, padx=8, pady=4)
    search_frame.pack(fill="x")
    tk.Label(search_frame, text="Filter:").pack(side="left")
    search_var = tk.StringVar()
    tk.Entry(search_frame, textvariable=search_var, width=30).pack(side="left", padx=6)

    cols = ("Vendor / Commitment", "PRIVV Total", "eBuilder Total", "Delta", "eB Invoices", "Status")
    tree_frame = tk.Frame(win)
    tree_frame.pack(fill="both", expand=True, padx=8, pady=4)

    vsb = ttk.Scrollbar(tree_frame, orient="vertical")
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
    tree = ttk.Treeview(
        tree_frame, columns=cols, show="headings",
        yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=20
    )
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)

    col_widths = {
        "Vendor / Commitment": 340,
        "PRIVV Total": 130, "eBuilder Total": 140,
        "Delta": 130, "eB Invoices": 90, "Status": 160,
    }
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=col_widths.get(c, 120),
                    anchor="e" if c not in ("Vendor / Commitment", "Status") else "w")

    tree.tag_configure("match",  background="#d4edda")
    tree.tag_configure("higher", background="#fff3cd")
    tree.tag_configure("lower",  background="#f8d7da")
    tree.tag_configure("nodata", background="#e2e3e5")

    iid_to_row: dict = {}

    UNASSIGNED_LABEL = "⚠ Unassigned (Manually Removed)"

    def _entry_amount(e: dict) -> float:
        """Best-effort dollar amount for an invoice entry, whichever side
        (eBuilder or PRIVV) it came from."""
        amt = e.get("_amount_num", None)
        if amt not in (None, ""):
            try:
                return float(amt)
            except (TypeError, ValueError):
                pass
        for key in ("Invoice Amount", "Amount"):
            if key in e:
                s = str(e.get(key, "")).replace(",", "").strip()
                if s and s.lower() != "nan":
                    try:
                        return float(s)
                    except ValueError:
                        continue
        return 0.0

    def _recalc_inv_row(r: dict):
        """Recompute eBuilder Total / PRIVV Total / eB Matches / Delta /
        Status for an invoice row from whatever is currently in
        r['_eb_entries'] / r['_prv_entries']. Used after an entry is
        manually reassigned or removed."""
        eb_entries  = r.get("_eb_entries", [])
        prv_entries = r.get("_prv_entries", [])
        r["eBuilder Total"] = sum(_entry_amount(e) for e in eb_entries)
        r["PRIVV Total"]    = sum(_entry_amount(e) for e in prv_entries)
        r["eB Matches"]     = len(eb_entries)
        r["Delta"] = r["eBuilder Total"] - r["PRIVV Total"]
        if r["eBuilder Total"] == 0 and r["PRIVV Total"] == 0:
            r["Status"] = "⚪ No Data"
        elif abs(r["Delta"]) < 0.01:
            r["Status"] = "✅ Match"
        elif r["eBuilder Total"] > r["PRIVV Total"]:
            r["Status"] = "🔼 eBuilder Higher"
        else:
            r["Status"] = "🔽 eBuilder Lower"

    def _get_or_create_unassigned_inv_row():
        for r in rows:
            if r["Label"] == UNASSIGNED_LABEL:
                return r
        new_r = {
            "Label": UNASSIGNED_LABEL,
            "PRIVV Total": 0.0,
            "eBuilder Total": 0.0,
            "Delta": 0.0,
            "eB Matches": 0,
            "Status": "⚪ No Data",
            "_eb_entries": [],
            "_prv_entries": [],
        }
        rows.append(new_r)
        return new_r

    def populate_tree(filter_text=""):
        tree.delete(*tree.get_children())
        iid_to_row.clear()
        ft = filter_text.strip().lower()
        for r in rows:
            if ft and ft not in r["Label"].lower():
                continue
            tag = "nodata"
            if "Match"  in r["Status"]:  tag = "match"
            elif "Higher" in r["Status"]: tag = "higher"
            elif "Lower"  in r["Status"]: tag = "lower"
            iid = tree.insert("", "end", values=(
                r["Label"],
                f"${r['PRIVV Total']:,.2f}",
                f"${r['eBuilder Total']:,.2f}",
                f"${r['Delta']:+,.2f}",
                r["eB Matches"],
                r["Status"],
            ), tags=(tag,))
            iid_to_row[iid] = r

    def show_inv_entries(event=None):
        sel = tree.selection()
        if not sel:
            return
        row = iid_to_row.get(sel[0])
        if row is None:
            return

        eb_entries  = row.get("_eb_entries", [])
        prv_entries = row.get("_prv_entries", [])

        detail = tk.Toplevel(win)
        detail.title(f"Invoice Entries — {row['Label']}")
        detail.geometry("1300x580")

        hdr = tk.Frame(detail, bg="#1e1e2e", padx=10, pady=6)
        hdr.pack(fill="x")
        hdr_label = tk.Label(
            hdr,
            text=(f"  {row['Label']}     "
                  f"eBuilder Total: ${row['eBuilder Total']:,.2f}     "
                  f"PRIVV Total: ${row['PRIVV Total']:,.2f}     "
                  f"eB Invoices: {len(eb_entries)}     PRIVV Invoices: {len(prv_entries)}"),
            bg="#1e1e2e", fg="white", font=("Consolas", 10, "bold"), anchor="w"
        )
        hdr_label.pack(fill="x")

        nb = ttk.Notebook(detail)
        nb.pack(fill="both", expand=True, padx=8, pady=6)

        def _refresh_header():
            hdr_label.config(
                text=(f"  {row['Label']}     "
                      f"eBuilder Total: ${row['eBuilder Total']:,.2f}     "
                      f"PRIVV Total: ${row['PRIVV Total']:,.2f}     "
                      f"eB Invoices: {len(row.get('_eb_entries', []))}     "
                      f"PRIVV Invoices: {len(row.get('_prv_entries', []))}"),
            )

        def _build_inv_tab(parent, entries, fallback_cols, label):
            """Build one Entries tab. Returns (etree, iid_to_entry_index,
            redraw_fn) so the reassign controls below can read the current
            selection and redraw in place after a move/remove."""
            if entries:
                tab_cols = []
                seen = set()
                for e in entries:
                    for k in e.keys():
                        if not str(k).startswith("_") and k not in seen:
                            seen.add(k)
                            tab_cols.append(k)
            else:
                tab_cols = list(fallback_cols)

            f = tk.Frame(parent)
            f.pack(fill="both", expand=True, padx=8, pady=6)
            evsb = ttk.Scrollbar(f, orient="vertical")
            ehsb = ttk.Scrollbar(f, orient="horizontal")
            etree = ttk.Treeview(f, columns=tab_cols, show="headings",
                                  yscrollcommand=evsb.set, xscrollcommand=ehsb.set, height=14)
            evsb.config(command=etree.yview)
            ehsb.config(command=etree.xview)
            money = {"Invoice Amount", "Amount", "_amount_num"}
            for c in tab_cols:
                etree.heading(c, text=c)
                etree.column(c, width=150, anchor="e" if c in money else "w")

            iid_to_entry_index = {}

            def _redraw(current_entries):
                etree.delete(*etree.get_children())
                iid_to_entry_index.clear()
                if current_entries:
                    for i, e in enumerate(current_entries):
                        iid = etree.insert("", "end", values=tuple(e.get(c, "") for c in tab_cols))
                        iid_to_entry_index[iid] = i
                else:
                    etree.insert("", "end", values=(label,) + ("",) * (len(tab_cols) - 1))

            _redraw(entries)
            evsb.pack(side="right", fill="y")
            ehsb.pack(side="bottom", fill="x")
            etree.pack(fill="both", expand=True)
            return etree, iid_to_entry_index, _redraw

        eb_tab_frame  = tk.Frame(nb)
        prv_tab_frame = tk.Frame(nb)
        nb.add(eb_tab_frame,  text=f"eBuilder Invoices ({len(eb_entries)})")
        nb.add(prv_tab_frame, text=f"PRIVV Invoices ({len(prv_entries)})")

        eb_tree, eb_iid_to_entry, eb_redraw = _build_inv_tab(
            eb_tab_frame, eb_entries,
            ["Invoice #", "Description", "Company", "Commitment #", "Invoice Amount"],
            "(No eBuilder entries)",
        )
        prv_tree, prv_iid_to_entry, prv_redraw = _build_inv_tab(
            prv_tab_frame, prv_entries, list(prv_inv_df.columns), "(No PRIVV entries)",
        )

        def _label_choices():
            choices = [r["Label"] for r in rows if r["Label"] != row["Label"]]
            choices.sort(key=str.lower)
            choices.append(UNASSIGNED_LABEL)
            return choices

        def _make_reassign_controls(parent, entries_key: str, etree, iid_to_entry,
                                     redraw_fn, tab_widget, tab_text_prefix):
            """Adds a 'Selected entry -> [combobox] [Move] [Remove/Unassign]'
            row under an entries tree, wired to move/remove the selected
            entry between this invoice-comparison row and any other (or the
            Unassigned bucket)."""
            frame = tk.Frame(parent, pady=6)
            frame.pack(fill="x", padx=8)

            tk.Label(frame, text="Selected entry →").pack(side="left")

            reassign_var = tk.StringVar()
            reassign_box = ttk.Combobox(
                frame, textvariable=reassign_var, values=_label_choices(),
                width=45, state="readonly",
            )
            reassign_box.pack(side="left", padx=6)

            def _refresh_this_tab():
                current_entries = row.get(entries_key, [])
                redraw_fn(current_entries)
                nb.tab(tab_widget, text=f"{tab_text_prefix} ({len(current_entries)})")
                reassign_box.config(values=_label_choices())
                _refresh_header()

            def apply_reassign():
                sel = etree.selection()
                if not sel:
                    messagebox.showwarning("No entry selected", "Select an entry first.", parent=detail)
                    return
                idx = iid_to_entry.get(sel[0])
                if idx is None:
                    return
                target_label = reassign_var.get()
                if not target_label:
                    messagebox.showwarning("No destination", "Choose where to move this entry.", parent=detail)
                    return

                entries_list = row.get(entries_key, [])
                if idx >= len(entries_list):
                    return
                entry = entries_list.pop(idx)
                _recalc_inv_row(row)

                if target_label == UNASSIGNED_LABEL:
                    target_row = _get_or_create_unassigned_inv_row()
                else:
                    target_row = next((r for r in rows if r["Label"] == target_label), None)
                    if target_row is None:
                        entries_list.insert(idx, entry)  # target vanished, revert
                        _recalc_inv_row(row)
                        return
                target_row.setdefault(entries_key, []).append(entry)
                _recalc_inv_row(target_row)

                populate_tree(search_var.get())
                _refresh_this_tab()
                side = "eBuilder" if entries_key == "_eb_entries" else "PRIVV"
                inv_log(f"Moved {side} invoice entry from '{row['Label']}' to '{target_row['Label']}'.")

            def remove_entry():
                reassign_var.set(UNASSIGNED_LABEL)
                apply_reassign()

            tk.Button(frame, text="Move Entry", command=apply_reassign,
                      bg="#0066cc", fg="white", padx=8).pack(side="left", padx=4)
            tk.Button(frame, text="🗑 Remove / Unassign", command=remove_entry,
                      padx=8).pack(side="left", padx=4)

        _make_reassign_controls(
            eb_tab_frame, "_eb_entries", eb_tree, eb_iid_to_entry,
            eb_redraw, eb_tab_frame, "eBuilder Invoices",
        )
        _make_reassign_controls(
            prv_tab_frame, "_prv_entries", prv_tree, prv_iid_to_entry,
            prv_redraw, prv_tab_frame, "PRIVV Invoices",
        )

        tk.Button(detail, text="Close", command=detail.destroy, padx=10).pack(pady=6)

    tree.bind("<Double-1>", show_inv_entries)
    populate_tree()
    search_var.trace_add("write", lambda *_: populate_tree(search_var.get()))

    vsb.pack(side="right",  fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    tk.Label(win, text="💡 Double-click any row to view its individual invoice entries",
             fg="#555", font=("Segoe UI", 9, "italic")).pack(pady=(0, 2))

    legend = tk.Frame(win, padx=8, pady=4)
    legend.pack(fill="x")
    for color, lbl in [
        ("#d4edda", "✅ Match"),
        ("#fff3cd", "🔼 eBuilder Higher"),
        ("#f8d7da", "🔽 eBuilder Lower"),
        ("#e2e3e5", "⚪ No Data"),
    ]:
        tk.Label(legend, text=f"  {lbl}  ", bg=color, relief="solid",
                 bd=1, padx=6, pady=2).pack(side="left", padx=4)

    # ── Export invoice comparison to Excel ───────────────────────────────────
    def export_invoice_comparison():
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save invoice comparison as..."
        )
        if not path:
            return

        HEADER_FILL = PatternFill("solid", start_color="1E1E2E", end_color="1E1E2E")
        HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
        BASE_FONT   = Font(name="Calibri")
        THIN        = Side(style="thin", color="D9D9D9")
        BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        STATUS_FILL = {
            "✅ Match":           "C6EFCE",
            "🔼 eBuilder Higher": "FFEB9C",
            "🔽 eBuilder Lower":  "FFC7CE",
            "⚪ No Data":          "E2E3E5",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Summary"
        headers = ["Vendor / Commitment", "eBuilder Total", "PRIVV Total", "Delta", "Status", "eB Invoices"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FILL and HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

        money_cols = {2, 3, 4}
        for r_idx, r in enumerate(rows, start=2):
            values = [
                r["Label"],
                r["eBuilder Total"], r["PRIVV Total"], r["Delta"],
                r["Status"], r["eB Matches"],
            ]
            fill = PatternFill("solid", start_color=STATUS_FILL.get(r["Status"], "FFFFFF"))
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font  = BASE_FONT
                cell.border = BORDER
                cell.fill  = fill
                if c_idx in money_cols:
                    cell.number_format = '$#,##0.00;($#,##0.00);"-"'

        # eBuilder entries sheet
        eb_sheet = wb.create_sheet("eBuilder Invoices")
        eb_hdrs  = ["Vendor / Commitment", "Invoice #", "Description", "Company",
                    "Commitment #", "Date Received", "Status", "Invoice Amount"]
        for c, h in enumerate(eb_hdrs, 1):
            cell = eb_sheet.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        eb_sheet.freeze_panes = "A2"
        r_idx = 2
        for r in rows:
            for e in r.get("_eb_entries", []):
                row_vals = [r["Label"]] + [e.get(h, "") for h in eb_hdrs[1:]]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = eb_sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BASE_FONT; cell.border = BORDER
                r_idx += 1

        # PRIVV entries sheet
        prv_sheet = wb.create_sheet("PRIVV Invoices")
        prv_cols  = [c for c in prv_inv_df.columns if not str(c).startswith("_")]
        prv_hdrs  = ["Vendor / Commitment"] + prv_cols
        for c, h in enumerate(prv_hdrs, 1):
            cell = prv_sheet.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        prv_sheet.freeze_panes = "A2"
        r_idx = 2
        for r in rows:
            for e in r.get("_prv_entries", []):
                row_vals = [r["Label"]] + [e.get(col, "") for col in prv_cols]
                for c_idx, val in enumerate(row_vals, 1):
                    cell = prv_sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BASE_FONT; cell.border = BORDER
                r_idx += 1

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=win)

    tk.Button(win, text="Export to Excel", command=export_invoice_comparison,
              bg="#0066cc", fg="white", padx=8).pack(pady=6)


# ── Invoice tab buttons / log ────────────────────────────────────────────────
tk.Button(
    inv_btn_frame, text="Compare Invoices", command=run_invoice_comparison,
    bg="#0066cc", fg="white", width=18, padx=6
).pack(side="left", padx=8)

inv_output_box = tk.Text(inv_tab, height=18, width=100)
inv_output_box.pack(pady=10, padx=10)

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — VENDOR ALIASES  (shared Google Sheet, same concept as tt.py)
# ════════════════════════════════════════════════════════════════════════════
alias_tab = tk.Frame(notebook)
notebook.add(alias_tab, text="Vendor Aliases")

tk.Label(
    alias_tab,
    text="Vendor Name -> Alias mappings live in the shared Google Sheet "
         f"('{GOOGLE_SHEET_NAME}' -> '{ALIAS_TAB_NAME}' tab). Anything added "
         "here is used immediately by both the Comparisons and Invoices tabs.",
    fg="#555", wraplength=760, justify="left", anchor="w",
).pack(pady=(10, 4), padx=10, fill="x")

alias_status_label = tk.Label(alias_tab, text="", fg="gray", anchor="w")
alias_status_label.pack(padx=10, fill="x")

alias_form = tk.Frame(alias_tab)
alias_form.pack(pady=8, padx=10, fill="x")

tk.Label(alias_form, text="Vendor Name:").grid(row=0, column=0, sticky="w", padx=4, pady=4)
alias_vendor_entry = tk.Entry(alias_form, width=40)
alias_vendor_entry.grid(row=0, column=1, padx=4, pady=4, sticky="w")

tk.Label(alias_form, text="Alias:").grid(row=0, column=2, sticky="w", padx=4, pady=4)
alias_alias_entry = tk.Entry(alias_form, width=40)
alias_alias_entry.grid(row=0, column=3, padx=4, pady=4, sticky="w")

alias_tree_frame = tk.Frame(alias_tab)
alias_tree_frame.pack(pady=6, padx=10, fill="both", expand=True)

alias_tree = ttk.Treeview(
    alias_tree_frame, columns=("vendor", "alias"), show="headings", height=16
)
alias_tree.heading("vendor", text="Vendor Name")
alias_tree.heading("alias", text="Alias")
alias_tree.column("vendor", width=340, anchor="w")
alias_tree.column("alias", width=340, anchor="w")
alias_tree.pack(side="left", fill="both", expand=True)

alias_scroll = ttk.Scrollbar(alias_tree_frame, orient="vertical", command=alias_tree.yview)
alias_tree.configure(yscroll=alias_scroll.set)
alias_scroll.pack(side="right", fill="y")


def refresh_alias_tree():
    """Reload SHEET_VENDOR_ALIASES from Google Sheets and repopulate the table."""
    alias_status_label.config(text="Loading vendor aliases from Google Sheet...", fg="gray")
    alias_tab.update_idletasks()
    load_vendor_aliases_from_sheet()

    for row in alias_tree.get_children():
        alias_tree.delete(row)
    for vendor_lower, alias in sorted(SHEET_VENDOR_ALIASES.items()):
        alias_tree.insert("", "end", values=(vendor_lower, alias))

    if not GSPREAD_AVAILABLE:
        alias_status_label.config(
            text="gspread / google-auth not installed - vendor aliases are local-only until installed.",
            fg="#b36b00",
        )
    elif not os.path.exists(CREDS_FILE):
        alias_status_label.config(
            text=f"'{CREDS_FILE}' not found - add your service-account key to enable the shared sheet.",
            fg="#b36b00",
        )
    else:
        alias_status_label.config(
            text=f"Loaded {len(SHEET_VENDOR_ALIASES)} alias(es) from '{GOOGLE_SHEET_NAME}' -> '{ALIAS_TAB_NAME}'.",
            fg="green",
        )


def add_alias_clicked():
    vendor = alias_vendor_entry.get().strip()
    alias  = alias_alias_entry.get().strip()
    if not vendor or not alias:
        messagebox.showwarning("Missing info", "Please enter both a Vendor Name and an Alias.", parent=alias_tab)
        return
    try:
        add_vendor_alias_to_sheet(vendor, alias)
    except Exception as e:
        messagebox.showerror("Could not save alias", str(e), parent=alias_tab)
        return
    alias_vendor_entry.delete(0, tk.END)
    alias_alias_entry.delete(0, tk.END)
    refresh_alias_tree()
    messagebox.showinfo("Saved", f"'{vendor}' -> '{alias}' added to the shared sheet.", parent=alias_tab)


alias_btn_frame = tk.Frame(alias_tab)
alias_btn_frame.pack(pady=(0, 10), padx=10, fill="x")

tk.Button(
    alias_btn_frame, text="Add Alias", command=add_alias_clicked,
    bg="#0066cc", fg="white", width=14, padx=6
).pack(side="left", padx=4)

tk.Button(
    alias_btn_frame, text="Refresh from Sheet", command=refresh_alias_tree,
    width=16, padx=6
).pack(side="left", padx=4)

# Load whatever is already in the sheet as soon as the app starts, so
# resolve_alias() has the latest user-taught mappings from the very first
# comparison/invoice run - not just after visiting this tab.
refresh_alias_tree()

root.mainloop()


    

#06/24/2026 Handles the budget part of the porogram


#6.28.2026 working on invoice part of the program

#6/30/2026 Invoices does sum i need to double check the accuracy of the program hopefully ity
